from __future__ import annotations

import asyncio
import logging
import os
import random
import sqlite3
from datetime import datetime, date, time as dtime, timedelta, timezone
from typing import Optional
import pytz
import openpyxl  # For XLSX support
import html

# word1 features removed: duel, hunt, share, progress

from backup_restore import create_full_backup, list_backups, restore_full_backup, get_backup_size_info, create_user_data_backup

try:
    from math_telegram import MathBotHandler
    MATH_AVAILABLE = True
except ImportError:
    MATH_AVAILABLE = False
    log = logging.getLogger("quizbot")
    log.warning("Math module not available")

from telegram import (
    Update, InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton, Document
)
from telegram.ext import (
    Application, ApplicationBuilder, ContextTypes,
    CallbackQueryHandler, PollAnswerHandler, MessageHandler, CommandHandler, filters,
    JobQueue
)
from telegram.error import BadRequest, RetryAfter

from telegram.request import HTTPXRequest
from apscheduler.schedulers.background import BackgroundScheduler

# =====================
# Config
# =====================

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
if not BOT_TOKEN:
    raise ValueError("BOT_TOKEN environment variable is not set")

DB_PATH = os.getenv("DB_PATH", "/home/ubuntu/bot/bot.db")

# Parse ADMIN_IDS from environment variable (comma-separated)
ADMIN_IDS_ENV = os.getenv("ADMIN_IDS", "")
ADMIN_IDS = {int(uid.strip()) for uid in ADMIN_IDS_ENV.split(",") if uid.strip()} if ADMIN_IDS_ENV else set()

# global blitz sessiyalari: tg_id -> {active, correct, wrong, until, job}
BLITZ_SESSIONS: dict[int, dict] = {}

# global quiz sessions: tg_id -> {current_question_num, started_at, correct_count}
QUIZ_SESSIONS: dict[int, dict] = {}

# global math handler
MATH_HANDLER: Optional['MathBotHandler'] = None

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s | %(message)s",
)
log = logging.getLogger("quizbot")

# Timezone
TZ = pytz.timezone("Asia/Tashkent")
UTC = timezone.utc

# Points
POINTS_FOR_CORRECT = 5
POINTS_FOR_CORRECT_BLITZ = 7
POINTS_FOR_WRONG = -4
POINTS_FOR_ADDED = 0

# Review thresholds (customizable)
# number of correct answers required to promote a word to next review level
PROMOTE_THRESHOLD = 2
# number of wrong answers required to reset a word to level 0
RESET_THRESHOLD = 1

# =====================
# Multilang strings (improved English and Russian)
# =====================

LANGS = {
    "UZ": {
        "start": "Assalomu alaykum! 👋",
        "choose_lang": "🌍 Tilni tanlang:",
        "lang_uz": "🇺🇿 O‘zbekcha",
        "lang_ru": "🇷🇺 Русский",
        "lang_en": "🇬🇧 English",
        "menu_add": "➕ So‘z qo‘shish",
        "menu_quiz": "🎯 Viktorina",
        "menu_stats": "📊 Statistika",
        "menu_words": "🗒 Mening so‘zlarim",
        "menu_remind": "⏰ Eslatmalar",
        "menu_io": "📁 Import/Export",
        "menu_blitz": "⚡ Blitz",
        "menu_leader": "🏆 Reyting",
        "menu_lang": "🌐 Til",
        "menu_admin": "🛠 Admin",
        "menu_groups": "👥 Guruhlar",
        "menu_grammar": "📚 Grammatika",
        "grammar_files_list": "📚 Grammatika qoidalari:\n\nFaylni tanlang:",
        "menu_ielts": "🎓 IELTS",
        "ielts_files_list": "📚 IELTS Cambridge Kitoblari & Tinglash Testlari:\n\nFaylni tanlang:",
        "menu_math": "📐 Math",
        "menu_settings": "⚙️ Sozlamalar",
        "settings_panel": "⚙️ Sozlamalar\n\nProfil:\n{profile}\n\nQuiz sozlamalari:\n— So'z necha marta so'raladi: {quiz_repeat}\n— Noto'g'ri javoblardan keyin qayta boshlash: {restart_on_incorrect}",
        "profile_info": "Foydalanuvchi: @{username}\nTG ID: {tg_id}\nUID: {uid}",
        "settings_changed": "Sozlamalar yangilandi.",
        "send_quiz_repeat_prompt": "Iltimos, so'zni necha marta so'ralishini son sifatida yuboring (masalan: 2).",
        "send_restart_prompt": "Iltimos, qayta boshlash uchun noto'g'ri javoblar sonini yuboring (masalan: 3).",
        "invalid_number": "Noto'g'ri format. Iltimos butun son yuboring.",
        "add_prompt": "Yuboring: Inglizcha - Tarjima\nMasalan: Again - yana",
        "added_ok": "✅ Qo‘shildi: {eng} — {uz}",
        "format_error": "Format xato. Masalan: Again - yana",
        "no_words": "Hali so‘z yo‘q.",
        "quiz_no_words": "Avval so‘z qo‘shing: tugmadan foydalaning yoki “Inglizcha - Tarjima” yuboring.",
        "quiz_continue_prompt": "Oldingi viktorina sessiyangiz bor. Davom ettirishni xohlaysizmi?",
        "quiz_continue_yes": "✅ Ha, davom eting",
        "quiz_continue_no": "🔄 Yangi boshlash",
        "quiz_question_num": "Savol {num}",
        "import_export_text": "📁 Import/Export:\n— Export: ro‘yxatingizni XLSX qilib yuboraman.\n— Import: 2 ustunli XLSX yuboring (English,Uzbek).\nMisol:\nEnglish,Uzbek\nAgain,yana",
        "import_prompt": "📥 XLSX yuboring (English,Uzbek). Birinchi qatorda sarlavha bo‘lishi mumkin.",
        "import_done": "📥 Import tugadi. Qo‘shildi: {n} ta.",
        "reminder_panel": "⏰ Eslatmalar\nHolat: {state}\nMaqsad (kuniga): {goal}\nVaqt: {time} (Asia/Tashkent)",
        "reminder_msg_goal_reached": "👏 Zo‘r! Bugun {added}/{goal} ta so‘z qo‘shdingiz. Davom eting!",
        "reminder_msg_goal_left": "⏰ Eslatma: bugun maqsad {goal} ta. Hozircha {added}. Yana {left} ta qolmoqda.",
        "blitz_started": "⚡ Blitz boshlandi! Sizda {minutes} daqiqa vaqt bor. Har javobdan keyin yangi savol beriladi.",
        "blitz_time_up": "⏳ Blitz tugadi!\n✅ To‘g‘ri: {correct}\n❌ Noto‘g‘ri: {wrong}\n📊 Jami savol: {total}\n🏆 Ball: {score}",
        "blitz_choose_duration": "⏳ Blitz uchun vaqtni tanlang:",
        "blitz_duration_1": "1 daqiqa",
        "blitz_duration_3": "3 daqiqa",
        "blitz_duration_5": "5 daqiqa",
        "correct": "✅ To‘g‘ri!",
        "wrong": "❌ Noto‘g‘ri. To‘g‘ri javob: {ans}",
        "no_stats": "Hozircha ma'lumot yo'q.",
        "leader_choose": "🏆 Reyting turini tanlang:",
        "leader_none": "❌ Hozircha hech kim ball to‘plamagan.",
        "group_select_prompt": "Viktorina uchun guruhni tanlang:",
        "group_view_prompt": "So'zlarini ko'rish uchun guruhni tanlang:",
        "group_rename_prompt": "Guruh uchun yangi nom yuboring:",
        "group_rename_success": "✅ Guruh nomi o'zgartirildi.",
        "group_rename_fail": "❌ Guruh nomi o'zgartirilmadi (egalik yoki xato).",
        "group_delete_confirm": "Guruhni o'chirishni tasdiqlang:",
        "group_delete_yes": "🗑️ Ha, o'chirish",
        "group_delete_no": "❌ Yo'q",
        "group_delete_success": "✅ Guruh o'chirildi.",
        "group_delete_fail": "❌ Guruh o'chirilmadi.",
        "not_in_group": "Siz bu guruhda emassiz.",
        "group_add_prompt": "Guruhga so'z qo'shing: English - Tarjima",
        "group_add_success": "✅ Guruhga qo'shildi: {eng} — {uz}",
        "clear_all_confirm": "Hammasini o'chirishni tasdiqlang:",
        "clear_all_yes": "🗑️ Ha, hammasini o'chir",
        "clear_all_no": "❌ Yo'q",
        "clear_all_success": "✅ Hammasi o'chirildi.",
        "clear_all_fail": "❌ O'chirilmadi.",
        "admin_panel": "Admin panel:",
        "broadcast_prompt": "🔊 Broadcast matnini yuboring (faqat matn).",
        "cancelled": "Bekor qilindi.",
        "broadcast_sent": "Yuborildi ✅: {ok}\nXatolik ❌: {fail}",
        "users_list": "👥 Foydalanuvchilar ({start}–{end}/{total})",
        "user_search_prompt": "Foydalanuvchi @username yoki TG ID yuboring.",
        "invalid_format": "Noto‘g‘ri format.",
        "invalid_id": "ID noto‘g‘ri.",
        "not_found": "Topilmadi.",
        "operation_error": "Amal bajarilganda xatolik: {e}",
        "global_stats": "👥 Foydalanuvchilar: {users}\n🗒 So‘zlar: {words}\n➕ Qo‘shilgan: {added}\n✅ To‘g‘ri: {correct}\n❌ Xato: {wrong}",
        "export_preparing": "📦 XLSX fayllar tayyor. Yuborilyapti...",
        "groups_menu": "Guruhlar menyusi:",
        "groups_list": "👥 Guruhlar ({start}–{end}/{total})\nID        Nomi                  Yaratilgan\n--------------------------------------------------",
        "no_groups": "Guruhlar yo‘q.",
        "create_group_prompt": "Guruh nomini yuboring (masalan: Inglizcha A1)",
        "group_add_word_prompt": "Guruh ID va so'zni yuboring: group_id:english - uzbek\nMasalan: 1:Hello - Salom",
        "unknown_admin_action": "Noma'lum admin amali.",
        "correct_answer": "To‘g‘ri javob: {ans}",
        "language_changed": "Til o'zgartirildi.",
        "group_created": "Guruh yaratildi: {name} (ID: {gid})",
        "group_name_changed": "Guruh nomi o'zgartirildi.",
        "group_name_not_changed": "Guruh nomi o'zgartirilmadi (egalik yoki xato).",
        "group_deleted": "Guruh o'chirildi.",
        "group_not_deleted": "Guruh o'chirilmadi.",
        "word_added_to_group": "Guruhga so'z qo'shildi.",
        "no_permission": "Ruxsat yo'q.",
        "group_selected": "Guruh tanlandi: {name}",
        "personal_selected": "Barcha (shaxsiy) tanlandi.",
        "time_changed": "Vaqt o'zgartirildi.",
        "invalid_time_format": "Noto'g'ri format (HH:MM).",
        "points_edited": "Ballar tahrirlandi.",
        "word_added": "So'z qo'shildi.",
        "deleted": "O'chirildi.",
        "not_found_or_no_permission": "Topilmadi yoki ruxsat yo'q.",
        "filter_choose": "Filtrni tanlang:",
        "filter_all": "Hammasi",
        "filter_last_7_days": "Oxirgi 7 kun",
        "filter_last_30_days": "Oxirgi 30 kun",
        "back": "Orqaga",
        "delete_mode": "O'chirish rejimi (ushbu sahifa):",
        "awaiting_delete_number": "\n\nHozir o'chirmoqchi bo'lgan so'zning raqamini yozing (masalan: 1, 2, yoki 5)",
        "close": "Yopish",
        "import_cancelled": "Import bekor qilindi.",
        "group_import_cancelled": "Guruh importi bekor qilindi.",
        "no_words_in_group": "Guruhda so'zlar yo'q.",
        "enabled": "Yoqilgan",
        "disabled": "O'chirilgan",
        "custom_time_prompt": "HH:MM formatida vaqt yuboring (00:00 - 23:59, masalan, 14:30)",
        "multiple_added": "{count} ta so'z qo'shildi",
        "errors": "Xatolar:",
        "add_user_to_group_prompt": "Guruh ID va foydalanuvchi TG ID yuboring: group_id:user_tg_id",
        "user_added_to_group": "Foydalanuvchi guruhga qo'shildi.",
        "group_name_prompt_for_multi": "Bir nechta so'zlar uchun guruh nomi yuboring:",
        "group_add_word": "Guruhga so'z qo'shish",
        "group_edit_name": "Guruh nomini o'zgartirish",
        "group_delete": "Guruhni o'chirish",
        "group_add_user": "Guruhga foydalanuvchi qo'shish",
        # UZ bo'limiga qo'shing (masalan)
        "menu_duel": "⚔️ Duel",
        "menu_hunt": "🔎 So'z ovlash",
        "menu_share": "🔁 So'z almashish",
        "menu_daily_phrase": "📌 Kunlik ibora",
        "menu_motivation": "💬 Motivatsiya",
        "menu_progress": "📈 Progress",
        "admin_panel_heading": "Admin panel:",
        "all_personal_selected": "Hammasini tanlandi (shaxsiy).",
        "cancelled": "Bekor qilindi.",
        "choose_group_io": "Import/Export uchun guruhni tanlang yoki shaxsiy so'zlar:",
        "choose_group_delete": "O'chirish uchun guruhni tanlang:",
        "choose_group_rename": "Nomi o'zgartirish uchun guruhni tanlang:",
        "choose_month": "Oyni tanlang (YYYY-MM):",
        "goal_range_error": "Maqsad 1 dan 100 gacha bo'lishi kerak.",
        "groups_menu_heading": "Guruhlar menyusi:",
        "no_document": "Hech qanday hujjat topilmadi.",
        "xlsx_format_only": "Iltimos, faqat .xlsx formatidagi fayl yuboring (English, Uzbek).",
        "language_changed_success": "Til o'zgartirildi.",
        "no_data_in_month": "Oy {ym} da ma'lumot yo'q.",
        "send_goal_prompt": "Yilgan kunlik maqsadni yuboring (1-100 so'z):",
        "send_time_prompt": "Vaqtni HH:MM formatida yuboring (00:00 - 23:59, masalan, 14:30)",
        "text_not_found": "Matn topilmadi.",
        "unknown_command": "Noma'lum buyruq. Menyuni ishlatting.",
        "broadcast_confirm": "Broadcast: {txt}\nTasdiqlang:",
        "file_size_error": "Fayl hajmi {MAX_MB}MB dan katta — import bekor qilindi.",
        "file_read_error": "Faylni olishda xatolik: {e}",
        "xlsx_read_error": "XLSX ochilmadi: {e}",
        "no_group_owner": "Siz ushbu guruh egasi emassiz; import bekor qilindi.",
        "quiz_session_started": "Yangi viktorina boshlandi!",
        "choose_day": "Oy {ym} da kunni tanlang:",
        "continue_question": "Davom etilmoqda savol {question_num}",
        "no_words_found": "So'zlar topilmadi.",
        "word_added_success": "So'z qo'shildi.",
        "points_edited_success": "Ballar tahrirlandi.",
        "files_ready": "📦 XLSX fayllar tayyor. Yuborilyapti..."
    },
    "RU": {
        "start": "Привет! 👋",
        "choose_lang": "🌍 Выберите язык:",
        "lang_uz": "🇺🇿 Узбекский",
        "lang_ru": "🇷🇺 Русский",
        "lang_en": "🇬🇧 Английский",
        "menu_add": "➕ Добавить слово",
        "menu_quiz": "🎯 Викторина",
        "menu_stats": "📊 Статистика",
        "menu_words": "🗒 Мои слова",
        "menu_remind": "⏰ Напоминания",
        "menu_io": "📁 Импорт/Экспорт",
        "menu_blitz": "⚡ Блиц",
        "menu_leader": "🏆 Рейтинг",
        "menu_lang": "🌐 Язык",
        "menu_admin": "🛠 Админ",
        "menu_groups": "👥 Группы",
        "menu_grammar": "📚 Грамматика",
        "grammar_files_list": "📚 Правила грамматики:\n\nВыберите файл:",
        "menu_ielts": "🎓 IELTS",
        "ielts_files_list": "📚 IELTS Книги Cambridge & Тесты Аудирования:\n\nВыберите файл:",
        "menu_math": "📐 Math",
        "menu_settings": "⚙️ Настройки",
        "settings_panel": "⚙️ Настройки\n\nПрофиль:\n{profile}\n\nНастройки викторины:\n— Сколько раз спрашивать слово: {quiz_repeat}\n— После скольких неправильных ответов перезапуск: {restart_on_incorrect}",
        "profile_info": "Пользователь: @{username}\nTG ID: {tg_id}\nUID: {uid}",
        "settings_changed": "Настройки обновлены.",
        "send_quiz_repeat_prompt": "Отправьте число повторов слова (например: 2).",
        "send_restart_prompt": "Отправьте количество неправильных ответов для перезапуска (например: 3).",
        "invalid_number": "Неверный формат. Отправьте целое число.",
        "add_prompt": "Отправьте: English - Перевод\nНапример: Again - yana",
        "added_ok": "✅ Добавлено: {eng} — {uz}",
        "format_error": "Неверный формат. Например: Again - yana",
        "no_words": "Слов пока нет.",
        "quiz_no_words": "Сначала добавьте слова: используйте кнопку или отправьте “English - Перевод”.",
        "quiz_continue_prompt": "У вас есть предыдущая сессия викторины. Хотите продолжить?",
        "quiz_continue_yes": "✅ Да, продолжить",
        "quiz_continue_no": "🔄 Начать новую",
        "quiz_question_num": "Вопрос {num}",
        "import_export_text": "📁 Импорт/Экспорт:\n— Экспорт: я пришлю ваш список в XLSX.\n— Импорт: отправьте XLSX с 2 столбцами (English,Uzbek).\nПример:\nEnglish,Uzbek\nAgain,yana",
        "import_prompt": "📥 Отправьте XLSX (English,Uzbek). Заголовок в первой строке возможен.",
        "import_done": "📥 Импорт завершён. Добавлено: {n} слов.",
        "reminder_panel": "⏰ Напоминания\nСостояние: {state}\nЦель (в день): {goal}\nВремя: {time} (Asia/Tashkent)",
        "reminder_msg_goal_reached": "👏 Отлично! Сегодня добавлено {added}/{goal} слов. Продолжайте!",
        "reminder_msg_goal_left": "⏰ Напоминание: цель {goal} слов на сегодня. Пока {added}. Осталось {left}.",
        "blitz_started": "⚡ Блиц начат! У вас {minutes} минут. После каждого ответа — новый вопрос.",
        "blitz_time_up": "⏳ Блиц завершён!\n✅ Правильно: {correct}\n❌ Неправильно: {wrong}\n📊 Всего вопросов: {total}\n🏆 Очки: {score}",
        "blitz_choose_duration": "⏳ Выберите время для блица:",
        "blitz_duration_1": "1 минута",
        "blitz_duration_3": "3 минуты",
        "blitz_duration_5": "5 минут",
        "correct": "✅ Верно!",
        "wrong": "❌ Неверно. Правильный ответ: {ans}",
        "no_stats": "Пока нет данных.",
        "leader_choose": "🏆 Выберите период:",
        "leader_none": "❌ Пока никто не набрал очков.",
        "group_select_prompt": "Выберите группу для викторины:",
        "group_view_prompt": "Выберите группу для просмотра слов:",
        "group_rename_prompt": "Отправьте новое имя для группы:",
        "group_rename_success": "✅ Имя группы изменено.",
        "group_rename_fail": "❌ Имя группы не изменено (владение или ошибка).",
        "group_delete_confirm": "Подтвердите удаление группы:",
        "group_delete_yes": "🗑️ Да, удалить",
        "group_delete_no": "❌ Нет",
        "group_delete_success": "✅ Группа удалена.",
        "group_delete_fail": "❌ Группа не удалена.",
        "not_in_group": "Вы не в этой группе.",
        "group_add_prompt": "Добавьте слово в группу: English - Перевод",
        "group_add_success": "✅ Добавлено в группу: {eng} — {uz}",
        "clear_all_confirm": "Подтвердите очистку всех слов:",
        "clear_all_yes": "🗑️ Да, очистить все",
        "clear_all_no": "❌ Нет",
        "clear_all_success": "✅ Все очищено.",
        "clear_all_fail": "❌ Не очищено.",
        "admin_panel": "Админ панель:",
        "broadcast_prompt": "🔊 Отправьте текст broadcast (только текст).",
        "cancelled": "Отменено.",
        "broadcast_sent": "Отправлено ✅: {ok}\nОшибка ❌: {fail}",
        "users_list": "👥 Пользователи ({start}–{end}/{total})",
        "user_search_prompt": "Отправьте @username или TG ID пользователя.",
        "invalid_format": "Неверный формат.",
        "invalid_id": "ID неверный.",
        "not_found": "Не найдено.",
        "operation_error": "Ошибка при выполнении операции: {e}",
        "global_stats": "👥 Пользователи: {users}\n🗒 Слова: {words}\n➕ Добавлено: {added}\n✅ Верно: {correct}\n❌ Ошибка: {wrong}",
        "export_preparing": "📦 XLSX файлы готовы. Отправляются...",
        "groups_menu": "Меню групп:",
        "groups_list": "👥 Группы ({start}–{end}/{total})\nID        Имя                   Создано\n--------------------------------------------------",
        "no_groups": "Групп нет.",
        "create_group_prompt": "Отправьте имя группы (например: Английский A1)",
        "group_add_word_prompt": "Отправьте ID группы и слово: group_id:english - перевод\nПример: 1:Hello - Привет",
        "unknown_admin_action": "Неизвестное админ действие.",
        "correct_answer": "Правильный ответ: {ans}",
        "language_changed": "Язык изменен.",
        "group_created": "Группа создана: {name} (ID: {gid})",
        "group_name_changed": "Имя группы изменено.",
        "group_name_not_changed": "Имя группы не изменено (владение или ошибка).",
        "group_deleted": "Группа удалена.",
        "group_not_deleted": "Группа не удалена.",
        "word_added_to_group": "Слово добавлено в группу.",
        "no_permission": "Нет разрешения.",
        "group_selected": "Группа выбрана: {name}",
        "personal_selected": "Все (личные) выбрано.",
        "time_changed": "Время изменено.",
        "invalid_time_format": "Неверный формат (HH:MM).",
        "points_edited": "Очки отредактированы.",
        "word_added": "Слово добавлено.",
        "deleted": "Удалено.",
        "not_found_or_no_permission": "Не найдено или нет разрешения.",
        "filter_choose": "Выберите фильтр:",
        "filter_all": "Все",
        "filter_last_7_days": "Последние 7 дней",
        "filter_last_30_days": "Последние 30 дней",
        "back": "Назад",
        "delete_mode": "Режим удаления (эта страница):",
        "awaiting_delete_number": "\n\nТеперь напишите номер слова, которое хотите удалить (например: 1, 2 или 5)",
        "close": "Закрыть",
        "import_cancelled": "Импорт отменен.",
        "group_import_cancelled": "Импорт группы отменен.",
        "no_words_in_group": "В группе нет слов.",
        "enabled": "Включено",
        "disabled": "Выключено",
        "custom_time_prompt": "Отправьте время в формате HH:MM (00:00 - 23:59, например, 14:30)",
        "multiple_added": "{count} слов добавлено",
        "errors": "Ошибки:",
        "add_user_to_group_prompt": "Отправьте ID группы и TG ID пользователя: group_id:user_tg_id",
        "user_added_to_group": "Пользователь добавлен в группу.",
        "group_name_prompt_for_multi": "Отправьте имя группы для нескольких слов:",
        "group_add_word": "Добавить слово в группу",
        "group_edit_name": "Изменить имя группы",
        "group_delete": "Удалить группу",
        "group_add_user": "Добавить пользователя в группу",
        # RU bo'limiga qo'shing (masalan)
        "menu_duel": "⚔️ Дуэль",
        "menu_hunt": "🔎 Охота за словами",
        "menu_share": "🔁 Поделиться словом",
        "menu_daily_phrase": "📌 Фраза дня",
        "menu_motivation": "💬 Мотивация",
        "menu_progress": "📈 Прогресс",
        "admin_panel_heading": "Админ панель:",
        "all_personal_selected": "Все (личные) выбрано.",
        "cancelled": "Отменено.",
        "choose_group_io": "Выберите группу для импорта/экспорта или личные слова:",
        "choose_group_delete": "Выберите группу для удаления:",
        "choose_group_rename": "Выберите группу для переименования:",
        "choose_month": "Выберите месяц (YYYY-MM):",
        "goal_range_error": "Цель должна быть от 1 до 100.",
        "groups_menu_heading": "Меню групп:",
        "no_document": "Документ не найден.",
        "xlsx_format_only": "Пожалуйста, отправьте только файл .xlsx (English, Uzbek).",
        "language_changed_success": "Язык изменён.",
        "no_data_in_month": "Нет данных в месяце {ym}.",
        "send_goal_prompt": "Отправьте желаемую дневную цель (1-100 слов):",
        "send_time_prompt": "Отправьте время в формате HH:MM (00:00 - 23:59, например, 14:30)",
        "text_not_found": "Текст не найден.",
        "unknown_command": "Неизвестная команда. Используйте меню.",
        "broadcast_confirm": "Broadcast: {txt}\nПотвердить:",
        "file_size_error": "Размер файла больше {MAX_MB}MB — импорт отменён.",
        "file_read_error": "Ошибка при чтении файла: {e}",
        "xlsx_read_error": "XLSX не открыть: {e}",
        "no_group_owner": "Вы не владелец этой группы; импорт отменён.",
        "quiz_session_started": "Новая викторина начата!",
        "choose_day": "Выберите день в {ym}:",
        "continue_question": "Продолжается вопрос {question_num}",
        "no_words_found": "Слова не найдены.",
        "word_added_success": "Слово добавлено.",
        "points_edited_success": "Очки отредактированы.",
        "files_ready": "📦 XLSX файлы готовы. Отправляются...",
        "menu_daily_phrase": "📌 Фраза дня",
        "menu_motivation": "💬 Мотивация",
        "menu_progress": "📈 Прогресс"

    },
    "EN": {
        "start": "Hello! 👋",
        "choose_lang": "🌍 Choose language:",
        "lang_uz": "🇺🇿 Uzbek",
        "lang_ru": "🇷🇺 Russian",
        "lang_en": "🇬🇧 English",
        "menu_add": "➕ Add word",
        "menu_quiz": "🎯 Quiz",
        "menu_stats": "📊 Stats",
        "menu_words": "🗒 My words",
        "menu_remind": "⏰ Reminders",
        "menu_io": "📁 Import/Export",
        "menu_blitz": "⚡ Blitz",
        "menu_leader": "🏆 Leaderboard",
        "menu_lang": "🌐 Language",
        "menu_admin": "🛠 Admin",
        "menu_groups": "👥 Groups",
        "menu_grammar": "📚 Grammar",
        "grammar_files_list": "📚 Grammar Rules:\n\nSelect a file:",
        "menu_ielts": "🎓 IELTS",
        "ielts_files_list": "📚 IELTS Cambridge Books & Listening Tests:\n\nSelect a file:",
        "menu_math": "📐 Math",
        "menu_settings": "⚙️ Settings",
        "settings_panel": "⚙️ Settings\n\nProfile:\n{profile}\n\nQuiz settings:\n— Times to ask word: {quiz_repeat}\n— Restart after incorrect answers: {restart_on_incorrect}",
        "profile_info": "User: @{username}\nTG ID: {tg_id}\nUID: {uid}",
        "settings_changed": "Settings updated.",
        "send_quiz_repeat_prompt": "Please send how many times a word should be asked (e.g. 2).",
        "send_restart_prompt": "Please send number of incorrect answers before restart (e.g. 3).",
        "invalid_number": "Invalid format. Send an integer.",
        "add_prompt": "Send: English - Translation\nExample: Again - yana",
        "added_ok": "✅ Added: {eng} — {uz}",
        "format_error": "Wrong format. Example: Again - yana",
        "no_words": "No words yet.",
        "quiz_no_words": "Add words first: use button or send “English - Translation”.",
        "quiz_continue_prompt": "You have a previous quiz session. Do you want to continue?",
        "quiz_continue_yes": "✅ Yes, continue",
        "quiz_continue_no": "🔄 Start new",
        "quiz_question_num": "Question {num}",
        "import_export_text": "📁 Import/Export:\n— Export: I'll send your list as XLSX.\n— Import: send 2-column XLSX (English,Uzbek).\nExample:\nEnglish,Uzbek\nAgain,yana",
        "import_prompt": "📥 Send XLSX (English,Uzbek). Header in first row is allowed.",
        "import_done": "📥 Import finished. Added: {n} words.",
        "reminder_panel": "⏰ Reminders\nState: {state}\nGoal (per day): {goal}\nTime: {time} (Asia/Tashkent)",
        "reminder_msg_goal_reached": "👏 Great! Today {added}/{goal} words added. Keep going!",
        "reminder_msg_goal_left": "⏰ Reminder: goal {goal} words today. Now {added}. {left} left.",
        "blitz_started": "⚡ Blitz started! You have {minutes} minutes. Each answer gives a new question.",
        "blitz_time_up": "⏳ Blitz finished!\n✅ Correct: {correct}\n❌ Wrong: {wrong}\n📊 Total questions: {total}\n🏆 Score: {score}",
        "blitz_choose_duration": "⏳ Choose duration for blitz:",
        "blitz_duration_1": "1 minute",
        "blitz_duration_3": "3 minutes",
        "blitz_duration_5": "5 minutes",
        "correct": "✅ Correct!",
        "wrong": "❌ Wrong. Correct answer: {ans}",
        "no_stats": "No data yet.",
        "leader_choose": "🏆 Choose period:",
        "leader_none": "❌ No one has collected points yet.",
        "group_select_prompt": "Choose group for quiz:",
        "group_view_prompt": "Choose group to view words:",
        "group_rename_prompt": "Send new name for group:",
        "group_rename_success": "✅ Group name changed.",
        "group_rename_fail": "❌ Group name not changed (ownership or error).",
        "group_delete_confirm": "Confirm group deletion:",
        "group_delete_yes": "🗑️ Yes, delete",
        "group_delete_no": "❌ No",
        "group_delete_success": "✅ Group deleted.",
        "group_delete_fail": "❌ Group not deleted.",
        "not_in_group": "You are not in this group.",
        "group_add_prompt": "Add word to group: English - Translation",
        "group_add_success": "✅ Added to group: {eng} — {uz}",
        "clear_all_confirm": "Confirm clearing all words:",
        "clear_all_yes": "🗑️ Yes, clear all",
        "clear_all_no": "❌ No",
        "clear_all_success": "✅ All cleared.",
        "clear_all_fail": "❌ Not cleared.",
        "admin_panel": "Admin panel:",
        "broadcast_prompt": "🔊 Send broadcast text (text only).",
        "cancelled": "Cancelled.",
        "broadcast_sent": "Sent ✅: {ok}\nError ❌: {fail}",
        "users_list": "👥 Users ({start}–{end}/{total})",
        "user_search_prompt": "Send @username or TG ID of the user.",
        "invalid_format": "Invalid format.",
        "invalid_id": "ID invalid.",
        "not_found": "Not found.",
        "operation_error": "Error during operation: {e}",
        "global_stats": "👥 Users: {users}\n🗒 Words: {words}\n➕ Added: {added}\n✅ Correct: {correct}\n❌ Wrong: {wrong}",
        "export_preparing": "📦 XLSX files ready. Sending...",
        "groups_menu": "Groups menu:",
        "groups_list": "👥 Groups ({start}–{end}/{total})\nID        Name                  Created\n--------------------------------------------------",
        "no_groups": "No groups.",
        "create_group_prompt": "Send group name (e.g., English A1)",
        "group_add_word_prompt": "Send group ID and word: group_id:english - translation\nExample: 1:Hello - Hello",
        "unknown_admin_action": "Unknown admin action.",
        "correct_answer": "Correct answer: {ans}",
        "language_changed": "Language changed.",
        "group_created": "Group created: {name} (ID: {gid})",
        "group_name_changed": "Group name changed.",
        "group_name_not_changed": "Group name not changed (ownership or error).",
        "group_deleted": "Group deleted.",
        "group_not_deleted": "Group not deleted.",
        "word_added_to_group": "Word added to group.",
        "no_permission": "No permission.",
        "group_selected": "Group selected: {name}",
        "personal_selected": "All (personal) selected.",
        "time_changed": "Time changed.",
        "invalid_time_format": "Invalid format (HH:MM).",
        "points_edited": "Points edited.",
        "word_added": "Word added.",
        "deleted": "Deleted.",
        "not_found_or_no_permission": "Not found or no permission.",
        "filter_choose": "Choose filter:",
        "filter_all": "All",
        "filter_last_7_days": "Last 7 days",
        "filter_last_30_days": "Last 30 days",
        "back": "Back",
        "delete_mode": "Delete mode (this page):",
        "awaiting_delete_number": "\n\nNow write the number of the word you want to delete (for example: 1, 2, or 5)",
        "close": "Close",
        "import_cancelled": "Import cancelled.",
        "group_import_cancelled": "Group import cancelled.",
        "no_words_in_group": "No words in the group.",
        "enabled": "Enabled",
        "disabled": "Disabled",
        "custom_time_prompt": "Send time in HH:MM format (00:00 - 23:59, e.g., 14:30)",
        "multiple_added": "{count} words added",
        "errors": "Errors:",
        "add_user_to_group_prompt": "Send group ID and user TG ID: group_id:user_tg_id",
        "user_added_to_group": "User added to group.",
        "group_name_prompt_for_multi": "Send group name for multiple words:",
        "group_add_word": "Add word to group",
        "group_edit_name": "Edit group name",
        "group_delete": "Delete group",
        "group_add_user": "Add user to group",
        # EN bo'limiga qo'shing
        "menu_duel": "⚔️ Duel",
        "menu_hunt": "🔎 Word Hunt",
        "menu_share": "🔁 Share word",
        "menu_daily_phrase": "📌 Phrase of the day",
        "menu_motivation": "💬 Motivation",
        "menu_progress": "📈 Progress",
        "admin_panel_heading": "Admin panel:",
        "all_personal_selected": "All (personal) selected.",
        "cancelled": "Cancelled.",
        "choose_group_io": "Choose group for Import/Export or personal words:",
        "choose_group_delete": "Choose group to delete:",
        "choose_group_rename": "Choose group to rename:",
        "choose_month": "Choose month (YYYY-MM):",
        "goal_range_error": "Goal must be between 1 and 100.",
        "groups_menu_heading": "Groups menu:",
        "no_document": "No document found.",
        "xlsx_format_only": "Please send only .xlsx file (English, Uzbek).",
        "language_changed_success": "Language changed.",
        "no_data_in_month": "No data in month {ym}.",
        "send_goal_prompt": "Send desired daily goal (1-100 words):",
        "send_time_prompt": "Send time in HH:MM format (00:00 - 23:59, e.g., 14:30)",
        "text_not_found": "Text not found.",
        "unknown_command": "Unknown command. Use the menu.",
        "broadcast_confirm": "Broadcast: {txt}\nConfirm:",
        "file_size_error": "File size is larger than {MAX_MB}MB — import cancelled.",
        "file_read_error": "Error reading file: {e}",
        "xlsx_read_error": "Cannot open XLSX: {e}",
        "no_group_owner": "You are not the owner of this group; import cancelled.",
        "quiz_session_started": "New quiz session started!",
        "choose_day": "Choose day in {ym}:",
        "continue_question": "Continuing question {question_num}",
        "no_words_found": "No words found.",
        "word_added_success": "Word added.",
        "points_edited_success": "Points edited.",
        "files_ready": "📦 XLSX files ready. Sending..."
    }
}

def t_for(user_id: int, key: str, **kwargs) -> str:
    lang = get_ui_lang(user_id)
    string = LANGS.get(lang, LANGS["UZ"]).get(key, f"[{key}]")
    return string.format(**kwargs)

def build_main_keyboard(user_id: int) -> ReplyKeyboardMarkup:
    lang = get_ui_lang(user_id)
    L = LANGS.get(lang, LANGS["UZ"])
    kb = [
        [KeyboardButton(L["menu_add"]), KeyboardButton(L["menu_io"])],
        [KeyboardButton(L["menu_quiz"]), KeyboardButton(L["menu_blitz"])],
        [KeyboardButton(L["menu_words"]), KeyboardButton(L["menu_remind"])],
        [KeyboardButton(L["menu_stats"]), KeyboardButton(L["menu_leader"])],
        [KeyboardButton(L["menu_lang"]), KeyboardButton(L["menu_groups"])],
        [KeyboardButton(L["menu_grammar"]), KeyboardButton(L["menu_ielts"])],
        [KeyboardButton(L["menu_math"]), KeyboardButton(L["menu_settings"])],
    ]
    # **Qo'shimcha tugmalar**
    # Duel/Hunt/Share/Progress features removed — do not show buttons

    row = None
    with db() as conn:
        row = conn.execute("SELECT tg_id FROM users WHERE id=?", (user_id,)).fetchone()
    tg_id = row["tg_id"] if row else user_id

    if is_admin_by_db_id_or_static(tg_id):
        kb.append([KeyboardButton(L["menu_admin"])])

    return ReplyKeyboardMarkup(kb, resize_keyboard=True)

# DB helpers
# =====================

def db() -> sqlite3.Connection:
    if DB_PATH is None:
        raise ValueError("DB_PATH is not set")
    # Ensure parent directory exists to avoid sqlite failures
    parent = os.path.dirname(os.path.abspath(DB_PATH))
    if parent and not os.path.exists(parent):
        try:
            os.makedirs(parent, exist_ok=True)
        except Exception:
            pass
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def _ensure_column(conn: sqlite3.Connection, table: str, col: str, decl: str):
    try:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {decl}")
    except sqlite3.OperationalError:
        pass  # already exists

def init_db():
    with db() as conn:
        conn.executescript(
            """
            PRAGMA foreign_keys = ON;

            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tg_id INTEGER UNIQUE NOT NULL,
                username TEXT,
                first_seen TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                role TEXT NOT NULL DEFAULT 'user',
                points INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                created_at TEXT NOT NULL,
                owner_id INTEGER NOT NULL,
                FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS users_groups (
                user_id INTEGER NOT NULL,
                group_id INTEGER NOT NULL,
                PRIMARY KEY (user_id, group_id),
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(group_id) REFERENCES groups(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS words (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                group_id INTEGER,
                english TEXT NOT NULL,
                uzbek   TEXT NOT NULL,
                created_at TEXT NOT NULL,
                review_level INTEGER DEFAULT 0,
                next_review TEXT,
                correct_count INTEGER DEFAULT 0,
                last_correct_date TEXT,
                wrong_count INTEGER DEFAULT 0,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(group_id) REFERENCES groups(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS stats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                action TEXT NOT NULL CHECK (action IN ('added','correct','wrong')),
                word_id INTEGER,
                created_at TEXT NOT NULL,
                local_date TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(word_id) REFERENCES words(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS settings (
                user_id INTEGER PRIMARY KEY,
                daily_goal INTEGER NOT NULL DEFAULT 10,
                remind_time TEXT NOT NULL DEFAULT '18:00',
                remind_enabled INTEGER NOT NULL DEFAULT 0,
                ui_lang TEXT NOT NULL DEFAULT 'UZ',
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_stats_user_date ON stats(user_id, local_date);
            CREATE INDEX IF NOT EXISTS idx_words_user ON words(user_id);
            CREATE INDEX IF NOT EXISTS idx_words_group ON words(group_id);
            """
        )
        _ensure_column(conn, "users", "active", "INTEGER NOT NULL DEFAULT 1")
        _ensure_column(conn, "users", "role", "TEXT NOT NULL DEFAULT 'user'")
        _ensure_column(conn, "users", "points", "INTEGER NOT NULL DEFAULT 0")
        _ensure_column(conn, "words", "review_level", "INTEGER DEFAULT 0")
        _ensure_column(conn, "words", "next_review", "TEXT")
        _ensure_column(conn, "words", "correct_count", "INTEGER DEFAULT 0")
        _ensure_column(conn, "words", "last_correct_date", "TEXT")
        _ensure_column(conn, "words", "wrong_count", "INTEGER DEFAULT 0")
        _ensure_column(conn, "words", "group_id", "INTEGER")
        _ensure_column(conn, "settings", "ui_lang", "TEXT NOT NULL DEFAULT 'UZ'")
        _ensure_column(conn, "settings", "quiz_repeat", "INTEGER NOT NULL DEFAULT 1")
        _ensure_column(conn, "settings", "restart_on_incorrect", "INTEGER NOT NULL DEFAULT 3")
        _ensure_column(conn, "groups", "owner_id", "INTEGER NOT NULL")

# =====================
# Time helpers
# =====================

def now_tz() -> datetime:
    return datetime.now(TZ)

def local_date() -> str:
    return now_tz().date().isoformat()

def local_today() -> date:
    return now_tz().date()

# =====================
# User helpers (DB)
# =====================

def get_or_create_user(tg_id: int, username: Optional[str]) -> int:
    with db() as conn:
        row = conn.execute("SELECT id FROM users WHERE tg_id=?", (tg_id,)).fetchone()
        if row:
            conn.execute("UPDATE users SET username=? WHERE tg_id=?", (username or "", tg_id))
            return row["id"]
        now = datetime.now(UTC).isoformat(timespec="seconds")
        conn.execute("INSERT INTO users (tg_id, username, first_seen) VALUES (?,?,?)", (tg_id, username or "", now))
        uid = conn.execute("SELECT id FROM users WHERE tg_id=?", (tg_id,)).fetchone()["id"]
        conn.execute("INSERT OR REPLACE INTO settings (user_id, daily_goal, remind_time, remind_enabled, ui_lang) VALUES (?,?,?,?,?)",
                     (uid, 10, "18:00", 0, "UZ"))
        return uid

def get_user_row_by_tg(tg_id: int) -> Optional[sqlite3.Row]:
    with db() as conn:
        return conn.execute("SELECT * FROM users WHERE tg_id=?", (tg_id,)).fetchone()

def set_user_active(tg_id: int, active: bool):
    with db() as conn:
        conn.execute("UPDATE users SET active=? WHERE tg_id=?", (1 if active else 0, tg_id))

def set_user_role(tg_id: int, role: str):
    with db() as conn:
        conn.execute("UPDATE users SET role=? WHERE tg_id=?", (role, tg_id))

def is_banned(tg_id: int) -> bool:
    row = get_user_row_by_tg(tg_id)
    return bool(row and row["active"] == 0)

def db_role_is_admin(tg_id: int) -> bool:
    row = get_user_row_by_tg(tg_id)
    return bool(row and (row["role"] == "admin"))

def is_admin_by_db_id_or_static(user_id_or_tg: int) -> bool:
    try:
        if user_id_or_tg in ADMIN_IDS:
            log.info(f"Admin check: {user_id_or_tg} is in ADMIN_IDS")
            return True
    except Exception as e:
        log.warning(f"Admin ID check failed: {e}")
    row = get_user_row_by_tg(user_id_or_tg)
    if row and row["role"] == "admin":
        log.info(f"Admin check: {user_id_or_tg} has admin role in DB")
        return True
    log.info(f"Admin check: {user_id_or_tg} is not admin")
    return False

def is_admin(tg_id: int) -> bool:
    return (tg_id in ADMIN_IDS) or db_role_is_admin(tg_id)

# =====================
# Group helpers - Redesigned as personal categories for words
# =====================

def create_group(name: str, owner_id: int) -> int:
    with db() as conn:
        now = datetime.now(UTC).isoformat(timespec="seconds")
        cur = conn.execute("INSERT INTO groups (name, created_at, owner_id) VALUES (?,?,?)", (name, now, owner_id))
        gid = cur.lastrowid
        conn.execute("INSERT INTO users_groups (user_id, group_id) VALUES (?,?)", (owner_id, gid))  # Ownerni qo'sh
    GROUPS_CACHE.pop(owner_id, None)  # Invalidate cache for this user
    return gid

def rename_group(group_id: int, new_name: str, requester_id: int) -> bool:
    with db() as conn:
        row = conn.execute("SELECT owner_id FROM groups WHERE id=?", (group_id,)).fetchone()
        if not row or (row["owner_id"] != requester_id and not is_admin(requester_id)):
            return False
        conn.execute("UPDATE groups SET name=? WHERE id=?", (new_name, group_id))
    GROUPS_CACHE.pop(requester_id, None)  # Invalidate cache for this user
    return True

def delete_group(group_id: int, requester_id: int) -> bool:
    with db() as conn:
        row = conn.execute("SELECT owner_id FROM groups WHERE id=?", (group_id,)).fetchone()
        if not row or (row["owner_id"] != requester_id and not is_admin(requester_id)):
            return False
        conn.execute("DELETE FROM words WHERE group_id=?", (group_id,))
        conn.execute("DELETE FROM users_groups WHERE group_id=?", (group_id,))
        conn.execute("DELETE FROM groups WHERE id=?", (group_id,))
    WORDS_CACHE.clear()
    GROUPS_CACHE.clear()
    QUIZ_CACHE.clear()
    return True

def is_group_owner(user_id: int, group_id: int) -> bool:
    with db() as conn:
        row = conn.execute("SELECT owner_id FROM groups WHERE id=?", (group_id,)).fetchone()
    return bool(row and row["owner_id"] == user_id)

def is_group_member(user_id: int, group_id: int) -> bool:
    with db() as conn:
        row = conn.execute("SELECT * FROM users_groups WHERE user_id=? AND group_id=?", (user_id, group_id)).fetchone()
    return bool(row)

def add_user_to_group(user_id: int, group_id: int, requester_id: int) -> bool:
    if not (is_group_owner(requester_id, group_id) or is_admin(requester_id)):
        return False
    with db() as conn:
        conn.execute("INSERT OR IGNORE INTO users_groups (user_id, group_id) VALUES (?,?)", (user_id, group_id))
    return True

def delete_group_word(word_id: int, group_id: int, requester_id: int) -> bool:
    if not (is_group_owner(requester_id, group_id) or is_admin(requester_id)):
        return False
    with db() as conn:
            conn.execute("DELETE FROM words WHERE id=? AND group_id=?", (word_id, group_id))
    WORDS_CACHE.clear()
    return True

def get_user_groups(user_id: int) -> list[sqlite3.Row]:
    """Get user groups with caching to reduce DB hits."""
    if user_id not in GROUPS_CACHE:
        with db() as conn:
            GROUPS_CACHE[user_id] = conn.execute(
                "SELECT g.id, g.name FROM groups g JOIN users_groups ug ON g.id = ug.group_id WHERE ug.user_id=?",
                (user_id,)
            ).fetchall()
    return GROUPS_CACHE[user_id]

def count_groups() -> int:
    with db() as conn:
        (n,) = conn.execute("SELECT COUNT(*) FROM groups").fetchone()
    return int(n)

def fetch_groups_page(offset: int, limit: int = 50) -> list[sqlite3.Row]:
    with db() as conn:
        return conn.execute(
            "SELECT id, name, created_at FROM groups ORDER BY id DESC LIMIT ? OFFSET ?",
            (limit, offset)
        ).fetchall()

def groups_page_text(offset: int, total: int) -> str:
    rows = fetch_groups_page(offset)
    if not rows:
        return "No groups."
    start = offset + 1
    end = min(offset + 50, total)
    header = "👥 Groups (showing {}-{}/{}):\n```\nID     Name                  Created\n{}\n".format(start, end, total, ("-"*50))
    body = "\n".join("{:<10} {:<20} {}".format(r['id'], r['name'], r['created_at'][:10]) for r in rows)
    return "```\n{}{}```".format(header, body)

def groups_page_kb(offset: int, total: int) -> InlineKeyboardMarkup:
    buttons = []
    prev_off = max(offset - 50, 0)
    next_off = offset + 50 if offset + 50 < total else None
    row = []
    if offset > 0:
        row.append(InlineKeyboardButton("⬅️ Previous 50", callback_data=f"admin:groups:{prev_off}"))
    if next_off is not None:
        row.append(InlineKeyboardButton("Next 50 ➡️", callback_data=f"admin:groups:{next_off}"))
    buttons.append(row)
    buttons.append([InlineKeyboardButton("⬅️ Admin menu", callback_data="admin:main")])
    return InlineKeyboardMarkup(buttons)

# =====================
# Words & Stats operations
# =====================

WORDS_CACHE: dict[tuple[int, Optional[int]], list] = {}
# track last asked word per (db_user_id, group_id) to avoid immediate repeats
LAST_ASKED: dict[tuple[int, Optional[int]], int] = {}

# Caching for groups and quizzes to reduce DB hits
GROUPS_CACHE: dict[int, list] = {}  # user_id -> groups list
QUIZ_CACHE: dict[tuple[int, Optional[int]], list] = {}  # (user_id, group_id) -> quiz words

def add_word(user_id: int, english: str, uzbek: str, group_id: Optional[int] = None) -> int:
    with db() as conn:
        now = datetime.now(UTC).isoformat(timespec="seconds")
        cur = conn.execute(
            "INSERT INTO words (user_id, group_id, english, uzbek, created_at, review_level, next_review, correct_count, last_correct_date, wrong_count) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (user_id, group_id, english.strip(), uzbek.strip(), now, 0, local_date(), 0, local_date(), 0),
        )
        conn.execute("INSERT INTO stats (user_id, action, word_id, created_at, local_date) VALUES (?,?,?,?,?)",
                     (user_id, "added", cur.lastrowid, now, local_date()))
        conn.execute("UPDATE users SET points = MAX(points + ?, 0) WHERE id=?", (POINTS_FOR_ADDED, user_id))
    WORDS_CACHE.clear()
    return cur.lastrowid


def parse_word_line(line: str) -> tuple[str, str]:
    """Parse a single line into (english, uzbek).

    Accepts separators: '-', '–', '—', ':' and is tolerant to whitespace.
    Raises ValueError if parsing fails.
    """
    if not line or not line.strip():
        raise ValueError("empty line")
    normalized = line.replace("–", "-").replace("—", "-")
    # prefer dash separator
    if "-" in normalized:
        parts = normalized.split("-", 1)
    elif ":" in normalized:
        parts = normalized.split(":", 1)
    else:
        raise ValueError("no separator found")
    eng = str(parts[0]).strip()
    uz = str(parts[1]).strip()
    if not eng or not uz:
        raise ValueError("empty english or uzbek")
    return eng, uz


def add_words_from_lines(user_id: int, lines: list[str], group_id: Optional[int] = None, batch_size: int = 50) -> tuple[int, list[str]]:
    """Add multiple lines (each containing a pair) and return (added_count, errors).

    This centralizes parsing and uses `add_word` to insert. Processes in batches to avoid timeouts.
    """
    added = 0
    errors: list[str] = []
    
    for idx, line in enumerate(lines, start=1):
        try:
            if not line or not line.strip():
                continue
            eng, uz = parse_word_line(line)
            add_word(user_id, eng, uz, group_id=group_id)
            added += 1
        except Exception as e:
            errors.append(f"Line {idx}: {e}")
    return added, errors

def delete_word_if_owner(word_id: int, user_id: int) -> bool:
    with db() as conn:
        row = conn.execute("SELECT id FROM words WHERE id=? AND user_id=?", (word_id, user_id)).fetchone()
        if not row:
            return False
        conn.execute("DELETE FROM words WHERE id=?", (word_id,))
    WORDS_CACHE.clear()
    return True

def delete_all_words(user_id: int, group_id: Optional[int] = None) -> bool:
    with db() as conn:
        if group_id is not None:
            if not is_group_owner(user_id, group_id):
                return False
            conn.execute("DELETE FROM words WHERE group_id=?", (group_id,))
        else:
            conn.execute("DELETE FROM words WHERE user_id=?", (user_id,))
    WORDS_CACHE.clear()
    return True

def admin_delete_word(word_id: int) -> bool:
    with db() as conn:
        row = conn.execute("SELECT id FROM words WHERE id=?", (word_id,)).fetchone()
        if not row:
            return False
        conn.execute("DELETE FROM words WHERE id=?", (word_id,))
    WORDS_CACHE.clear()
    return True

def admin_add_word_to_user(target_user_id: int, english: str, uzbek: str) -> int:
    return add_word(target_user_id, english, uzbek)

def pick_user_word(user_id: int, group_id: Optional[int] = None) -> Optional[sqlite3.Row]:
    key = (user_id, group_id)
    today = local_date()
    if key not in WORDS_CACHE:
        with db() as conn:
            if group_id:
                WORDS_CACHE[key] = conn.execute(
                    "SELECT * FROM words WHERE group_id=? AND (next_review IS NULL OR next_review <= ?)",
                    (group_id, today)
                ).fetchall()
            else:
                WORDS_CACHE[key] = conn.execute(
                    "SELECT * FROM words WHERE user_id=? AND (next_review IS NULL OR next_review <= ?)",
                    (user_id, today)
                ).fetchall()
    words = WORDS_CACHE[key]
    if not words:
        return None
    # Avoid returning the same word consecutively for this user+group when possible
    last_id = LAST_ASKED.get(key)
    if last_id and len(words) > 1:
        candidates = [w for w in words if w["id"] != last_id]
        if candidates:
            return random.choice(candidates)
    return random.choice(words)

def get_distractors(user_id: int, correct_word_id: int, needed: int = 3, group_id: Optional[int] = None) -> list[str]:
    with db() as conn:
        if group_id:
            rows = conn.execute("SELECT uzbek FROM words WHERE group_id=? AND id<>? ORDER BY RANDOM() LIMIT ?", (group_id, correct_word_id, max(needed,0))).fetchall()
        else:
            rows = conn.execute("SELECT uzbek FROM words WHERE user_id=? AND id<>? ORDER BY RANDOM() LIMIT ?", (user_id, correct_word_id, max(needed,0))).fetchall()
    opts = [r["uzbek"] for r in rows]
    placeholders = ["nomuvofiq", "aniq emas", "bog'liq emas", "bilinmaydi"]
    while len(opts) < needed:
        opts.append(random.choice(placeholders))
    return opts[:needed]

def record_stat(user_id: int, action: str, word_id: Optional[int], is_blitz: bool = False, group_id: Optional[int] = None):
    with db() as conn:
        now = datetime.now(UTC).isoformat(timespec="seconds")
        d = local_date()
        conn.execute("INSERT INTO stats (user_id, action, word_id, created_at, local_date) VALUES (?,?,?,?,?)",
                    (user_id, action, word_id, now, d))
        if word_id:
            cur = conn.cursor()
            cur.execute("SELECT correct_count, last_correct_date, review_level, wrong_count FROM words WHERE id=?", (word_id,))
            w = cur.fetchone()
            if w:
                if action == 'correct':
                    new_count = (w["correct_count"] or 0) + 1
                    conn.execute("UPDATE words SET correct_count=?, last_correct_date=?, wrong_count=? WHERE id=?", (new_count, local_date(), 0, word_id))
                    if new_count == 2:
                        level = w["review_level"] or 0
                        if level == 0:
                            days_add = 1
                        elif level == 1:
                            days_add = 3
                        elif level == 2:
                            days_add = 10
                        else:
                            days_add = 30
                        next_date = (local_today() + timedelta(days=days_add)).isoformat()
                        new_level = level + 1
                        conn.execute("UPDATE words SET review_level=?, next_review=?, correct_count=0, wrong_count=0 WHERE id=?", (new_level, next_date, word_id))
                        WORDS_CACHE.clear()
                elif action == 'wrong':
                    new_wrong = (w["wrong_count"] or 0) + 1
                    conn.execute("UPDATE words SET wrong_count=? WHERE id=?", (new_wrong, word_id))
                    if new_wrong >= 1:
                        conn.execute("UPDATE words SET review_level=0, next_review=?, correct_count=0, wrong_count=0 WHERE id=?", (local_date(), word_id))
                        WORDS_CACHE.clear()
        delta = 0
        if action == "added":
            delta = POINTS_FOR_ADDED
        elif action == "correct":
            delta = POINTS_FOR_CORRECT_BLITZ if is_blitz else POINTS_FOR_CORRECT
        elif action == "wrong":
            delta = POINTS_FOR_WRONG
        if delta:
            conn.execute("UPDATE users SET points = MAX(points + ?, 0) WHERE id=?", (delta, user_id))

def set_user_points(user_id: int, points: int):
    with db() as conn:
        conn.execute("UPDATE users SET points = ? WHERE id=?", (points, user_id))

# =====================
# Reporting helpers
# =====================

def month_list_for_user(user_id: int) -> list[str]:
    with db() as conn:
        rows = conn.execute("SELECT DISTINCT substr(local_date,1,7) AS ym FROM stats WHERE user_id=? ORDER BY ym DESC", (user_id,)).fetchall()
    return [r["ym"] for r in rows]

def days_in_month_for_user(user_id: int, ym: str) -> list[str]:
    like = f"{ym}-%"
    with db() as conn:
        rows = conn.execute("SELECT DISTINCT local_date FROM stats WHERE user_id=? AND local_date LIKE ? ORDER BY local_date", (user_id, like)).fetchall()
    return [r["local_date"] for r in rows]

def day_counts(user_id: int, d: str) -> dict:
    with db() as conn:
        rows = conn.execute("SELECT action, COUNT(*) as c FROM stats WHERE user_id=? AND local_date=? GROUP BY action", (user_id, d)).fetchall()
    out = {"added": 0, "correct": 0, "wrong": 0}
    for r in rows:
        out[r["action"]] = r["c"]
    return out

def added_words_on(user_id: int, d: str, limit: int = 20) -> list[tuple[str, str]]:
    with db() as conn:
        rows = conn.execute(
            """
            SELECT w.english, w.uzbek
            FROM words w
            JOIN stats s ON s.word_id = w.id
            WHERE s.user_id=? AND s.local_date=? AND s.action='added'
            ORDER BY w.id DESC LIMIT ?
            """, (user_id, d, limit)
        ).fetchall()
    return [(r["english"], r["uzbek"]) for r in rows]

# ---- Admin DB helpers ----
def count_users(active_only: bool = True) -> int:
    with db() as conn:
        if active_only:
            (n,) = conn.execute("SELECT COUNT(*) FROM users WHERE active=1").fetchone()
        else:
            (n,) = conn.execute("SELECT COUNT(*) FROM users").fetchone()
    return int(n)

def count_words_all() -> int:
    with db() as conn:
        (n,) = conn.execute("SELECT COUNT(*) FROM words").fetchone()
    return int(n)

def count_stats_all(action: Optional[str] = None) -> int:
    with db() as conn:
        if action:
            (n,) = conn.execute("SELECT COUNT(*) FROM stats WHERE action=?", (action,)).fetchone()
        else:
            (n,) = conn.execute("SELECT COUNT(*) FROM stats").fetchone()
    return int(n)

def top_users_current_month(limit: int = 10) -> list[sqlite3.Row]:
    ym = local_date()[:7]
    with db() as conn:
        rows = conn.execute("""
            SELECT u.tg_id, u.username, COUNT(*) AS c
            FROM stats s
            JOIN users u ON u.id = s.user_id
            WHERE s.action='correct' AND substr(s.local_date,1,7)=?
            GROUP BY s.user_id
            ORDER BY c DESC
            LIMIT ?
        """, (ym, limit)).fetchall()
    return rows

def iter_all_tg_ids() -> list[int]:
    with db() as conn:
        rows = conn.execute("SELECT tg_id FROM users WHERE active=1").fetchall()
    return [r["tg_id"] for r in rows]

def fetch_users_page(offset: int, limit: int = 10, active_only: bool = True, sort_by: str = "id") -> list[sqlite3.Row]:
    if offset < 0:
        offset = 0
    sort_field = {"id": "id DESC", "points": "points DESC", "username": "username ASC", "first_seen": "first_seen DESC"}.get(sort_by, "id DESC")
    with db() as conn:
        if active_only:
            rows = conn.execute(f"SELECT id, tg_id, username, first_seen, active, role, points FROM users WHERE active=1 ORDER BY {sort_field} LIMIT ? OFFSET ?", (limit, offset)).fetchall()
        else:
            rows = conn.execute(f"SELECT id, tg_id, username, first_seen, active, role, points FROM users ORDER BY {sort_field} LIMIT ? OFFSET ?", (limit, offset)).fetchall()
    return rows

def users_page_text(offset: int, total: int, active_only: bool = True, sort_by: str = "id") -> str:
    rows = fetch_users_page(offset, active_only=active_only, sort_by=sort_by)
    if not rows:
        return "No users."
    start = offset + 1
    end = min(offset + 10, total)
    text = f"👥 Users ({start}–{end}/{total})\n"
    with db() as conn:
        for r in rows:
            uname = r['username'] or 'None'
            tg_id = r['tg_id']
            points = r['points']
            words_count = conn.execute("SELECT COUNT(*) FROM words WHERE user_id=?", (r["id"],)).fetchone()[0]
            active = "Active" if r['active'] else "Banned"
            role = r['role']
            text += f"• @{uname} ({tg_id}) - Points: {points}, Words: {words_count}, Role: {role}, Status: {active}\n"
    return text

def users_page_kb(offset: int, total: int, rows: list[sqlite3.Row], active_only: bool = True, sort_by: str = "id") -> InlineKeyboardMarkup:
    buttons = []
    for r in rows:
        uname = r['username'] or 'None'
        buttons.append([InlineKeyboardButton(f"@{uname} ({r['tg_id']})", callback_data=f"admin:user_info:{r['id']}")])
    row = []
    prev_off = max(offset - 10, 0)
    next_off = offset + 10 if offset + 10 < total else None
    if offset > 0:
        row.append(InlineKeyboardButton("⬅️ Previous", callback_data=f"admin:users:{prev_off}:{'active' if active_only else 'all'}:{sort_by}"))
    if next_off is not None:
        row.append(InlineKeyboardButton("Next ➡️", callback_data=f"admin:users:{next_off}:{'active' if active_only else 'all'}:{sort_by}"))
    buttons.append(row)
    buttons.append([
        InlineKeyboardButton("🔢 ID", callback_data=f"admin:users:{offset}:{'active' if active_only else 'all'}:id"),
        InlineKeyboardButton("🏆 Points", callback_data=f"admin:users:{offset}:{'active' if active_only else 'all'}:points"),
        InlineKeyboardButton("📛 Name", callback_data=f"admin:users:{offset}:{'active' if active_only else 'all'}:username"),
        InlineKeyboardButton("📅 Registered", callback_data=f"admin:users:{offset}:{'active' if active_only else 'all'}:first_seen")
    ])
    buttons.append([
        InlineKeyboardButton("🔍 Search", callback_data="admin:user_search"),
        InlineKeyboardButton("👥 Active/All", callback_data=f"admin:users:{offset}:{'all' if active_only else 'active'}:{sort_by}")
    ])
    buttons.append([InlineKeyboardButton("⬅️ Admin menu", callback_data="admin:main")])
    return InlineKeyboardMarkup(buttons)

def get_user_db_id_from_query(query: str) -> Optional[int]:
    with db() as conn:
        if query.isdigit():
            row = conn.execute("SELECT id FROM users WHERE tg_id=?", (int(query),)).fetchone()
        elif query.startswith("@"):
            row = conn.execute("SELECT id FROM users WHERE username=?", (query[1:],)).fetchone()
        else:
            return None
        return row["id"] if row else None

def get_user_info_text(user_db_id: int) -> str:
    with db() as conn:
        row = conn.execute("SELECT * FROM users WHERE id=?", (user_db_id,)).fetchone()
    if not row:
        return "Not found."
    username = row["username"] or "None"
    tg_id = row["tg_id"]
    active = "Active" if row["active"] else "Banned"
    role = row["role"]
    points = row["points"]
    return f"👤 @{username} (TG: {tg_id})\nFirst: {row['first_seen']}\nStatus: {active}\nRole: {role}\nPoints: {points}"

def get_user_info_kb(user_db_id: int) -> InlineKeyboardMarkup:
    with db() as conn:
        row = conn.execute("SELECT active, role FROM users WHERE id=?", (user_db_id,)).fetchone()
    active = row["active"]
    role = row["role"]
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🗒 View words", callback_data=f"admin:words:{user_db_id}:0")],
        [InlineKeyboardButton("➕ Add word", callback_data=f"admin:user:{user_db_id}:add_word")],
        [InlineKeyboardButton("🏆 Edit points", callback_data=f"admin:user:{user_db_id}:edit_points")],
        [InlineKeyboardButton("🚫 Ban" if active else "✅ Unban", callback_data=f"admin:user:{user_db_id}:toggle_active")],
        [InlineKeyboardButton("🛠 Admin" if role != "admin" else "👤 User", callback_data=f"admin:user:{user_db_id}:toggle_admin")],
        [InlineKeyboardButton("⬅️ Users", callback_data=f"admin:users:0:active:id")]
    ])

# ---- Settings (reminders & ui_lang) ----
def get_settings(user_id: int) -> dict:
    with db() as conn:
        r = conn.execute("SELECT daily_goal, remind_time, remind_enabled, ui_lang, quiz_repeat, restart_on_incorrect FROM settings WHERE user_id=?", (user_id,)).fetchone()
    if not r:
        return {"daily_goal": 10, "remind_time": "18:00", "remind_enabled": 0, "ui_lang": "UZ", "quiz_repeat": 1, "restart_on_incorrect": 3}
    d = dict(r)
    # ensure keys exist with defaults
    d.setdefault("quiz_repeat", 1)
    d.setdefault("restart_on_incorrect", 3)
    return d

def set_settings(user_id: int, **kwargs):
    fields = []
    vals = []
    for k, v in kwargs.items():
        fields.append(f"{k}=?")
        vals.append(v)
    if not fields:
        return
    with db() as conn:
        conn.execute(f"UPDATE settings SET {', '.join(fields)} WHERE user_id=?", (*vals, user_id))

def get_ui_lang(user_id: int) -> str:
    st = get_settings(user_id)
    return st.get("ui_lang", "UZ")

def set_ui_lang(user_id: int, lang: str):
    set_settings(user_id, ui_lang=lang)

# =====================
# In-memory poll/job state
# =====================
ACTIVE_POLLS: dict[str, dict] = {}

# =====================
# Utils for keyboards / t
# =====================

def language_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(LANGS["UZ"]["lang_uz"], callback_data="lang:UZ"),
         InlineKeyboardButton(LANGS["RU"]["lang_ru"], callback_data="lang:RU"),
         InlineKeyboardButton(LANGS["EN"]["lang_en"], callback_data="lang:EN")]
    ])

def reminder_kb(enabled: bool, goal: int, hhmm: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Enabled ✅" if enabled else "Disabled ❌", callback_data="rem:toggle")],
        [InlineKeyboardButton(f"Goal: {goal} words/day", callback_data="rem:goal:custom")],
        [InlineKeyboardButton("09:00", callback_data="rem:time:09:00"),
         InlineKeyboardButton("12:00", callback_data="rem:time:12:00"),
         InlineKeyboardButton("18:00", callback_data="rem:time:18:00"),
         InlineKeyboardButton("21:00", callback_data="rem:time:21:00")],
        [InlineKeyboardButton("Custom time", callback_data="rem:time:custom")],
        [InlineKeyboardButton("Close", callback_data="rem:close")]
    ])

def quiz_continue_kb(lang: str) -> InlineKeyboardMarkup:
    L = LANGS.get(lang, LANGS["UZ"])
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(L["quiz_continue_yes"], callback_data="quiz_continue:yes"),
         InlineKeyboardButton(L["quiz_continue_no"], callback_data="quiz_continue:no")]
    ])

def blitz_duration_kb(lang: str = "UZ") -> InlineKeyboardMarkup:
    L = LANGS.get(lang, LANGS["UZ"])
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(L["blitz_duration_1"], callback_data="blitz_start:1"),
         InlineKeyboardButton(L["blitz_duration_3"], callback_data="blitz_start:3"),
         InlineKeyboardButton(L["blitz_duration_5"], callback_data="blitz_start:5")],
        [InlineKeyboardButton("❌ Cancel", callback_data="blitz_start:cancel")]
    ])

# =====================
# Reminder scheduling
# =====================

def _parse_hhmm(hhmm: str) -> dtime:
    hh, mm = hhmm.split(":")
    return dtime(hour=int(hh), minute=int(mm), tzinfo=TZ)

async def reminder_job(ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = ctx.job.data["tg_id"]
    row = get_user_row_by_tg(tg_id)
    if not row or row["active"] == 0:
        return
    user_id = row["id"]
    st = get_settings(user_id)
    if not st.get("remind_enabled"):
        return
    today = local_date()
    cnts = day_counts(user_id, today)
    added_today = cnts.get("added", 0)
    goal = st.get("daily_goal", 10)
    if added_today >= goal:
        msg = LANGS.get(st.get("ui_lang", "UZ"))["reminder_msg_goal_reached"].format(added=added_today, goal=goal)
    else:
        left = goal - added_today
        msg = LANGS.get(st.get("ui_lang", "UZ"))["reminder_msg_goal_left"].format(goal=goal, added=added_today, left=left)
    try:
        await ctx.bot.send_message(chat_id=tg_id, text=msg)
    except Exception as e:
        log.warning("reminder send failed for %s: %s", tg_id, e)

def schedule_user_reminder(app: Application, tg_id: int, hhmm: str, enable: bool):
    name = f"rem_{tg_id}"
    for j in app.job_queue.get_jobs_by_name(name):
        j.schedule_removal()
    if enable:
        try:
            t = _parse_hhmm(hhmm)
            app.job_queue.run_daily(reminder_job, time=t, name=name, data={"tg_id": tg_id})
        except Exception as e:
            log.warning("schedule failed: %s", e)

def reschedule_all(app: Application):
    with db() as conn:
        rows = conn.execute("SELECT u.tg_id, s.remind_time FROM settings s JOIN users u ON u.id = s.user_id WHERE s.remind_enabled=1 AND u.active=1").fetchall()
    for r in rows:
        schedule_user_reminder(app, r["tg_id"], r["remind_time"], True)

# =====================
# Quiz (poll) functions
# =====================

async def send_quiz_poll(context: ContextTypes.DEFAULT_TYPE, chat_id: int, db_user_id: int, tg_user_id: int, group_id: Optional[int] = None, question_num: int = 1) -> bool:
    word = pick_user_word(db_user_id, group_id)
    if not word:
        await context.bot.send_message(chat_id, t_for(db_user_id, "quiz_no_words"))
        QUIZ_SESSIONS.pop(tg_user_id, None)
        return False

    # Track this word to avoid consecutive repeats
    LAST_ASKED[(db_user_id, group_id)] = word["id"]

    distractors = get_distractors(db_user_id, word["id"], 3, group_id=group_id)
    options = [word["uzbek"]] + distractors
    random.shuffle(options)
    correct_idx = options.index(word["uzbek"])

    L = LANGS.get(get_ui_lang(db_user_id), LANGS["UZ"])
    q = L["quiz_question_num"].format(num=question_num) + f"\n\nTarjimasini toping: {word['english']}"
    
    msg = await context.bot.send_poll(
        chat_id=chat_id,
        question=q,
        options=options,
        type="quiz",
        correct_option_id=correct_idx,
        is_anonymous=False,
        explanation=f"To‘g‘ri javob: {word['uzbek']}",
    )
    if msg.poll is None:
        return False

    if tg_user_id not in QUIZ_SESSIONS:
        QUIZ_SESSIONS[tg_user_id] = {
            "current_question_num": 1,
            "started_at": now_tz().isoformat(),
            "correct_count": 0
        }
    else:
        QUIZ_SESSIONS[tg_user_id]["current_question_num"] = question_num

    ACTIVE_POLLS[msg.poll.id] = {
        "chat_id": chat_id,
        "message_id": msg.message_id,
        "tg_user_id": tg_user_id,
        "db_user_id": db_user_id,
        "word_id": word["id"],
        "correct_idx": correct_idx,
        "group_id": group_id,
        "question_num": question_num,
        "is_blitz": False,
    }
    return True

async def quiz_continue_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    data = q.data.split(":")[1]

    if data == "yes":
        sess = QUIZ_SESSIONS.get(u.id, {})
        question_num = sess.get("current_question_num", 1)
        group_id = context.user_data.get("selected_group")
        ok = await send_quiz_poll(context, q.message.chat_id, uid, u.id, group_id=group_id, question_num=question_num)
        if ok:
            await q.edit_message_text(f"Davom etilmoqda savol {question_num}")
        else:
            QUIZ_SESSIONS.pop(u.id, None)
        await q.message.delete()
        return
    elif data == "no":
        QUIZ_SESSIONS[u.id] = {
            "current_question_num": 1,
            "started_at": now_tz().isoformat(),
            "correct_count": 0
        }
        group_id = context.user_data.get("selected_group")
        ok = await send_quiz_poll(context, q.message.chat_id, uid, u.id, group_id=group_id, question_num=1)
        if ok:
            await q.edit_message_text("Yangi viktorina boshlandi!")
        else:
            QUIZ_SESSIONS.pop(u.id, None)
        await q.message.delete()
        return

# =====================
# Blitz functions
# =====================

async def blitz_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    context.user_data["pending_action"] = "blitz"
    await select_group(update, context)
    return

async def blitz_start_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    data = q.data.split(":")[1]

    if data == "cancel":
        await q.message.delete()
        return

    try:
        minutes = int(data)
    except Exception:
        await q.edit_message_text("Invalid time.", reply_markup=None)
        return

    until = now_tz() + timedelta(minutes=minutes)
    job_name = f"blitz_{u.id}"
    job = context.application.job_queue.run_once(blitz_time_up_job, until - now_tz(), name=job_name, data={"tg_id": u.id})

    BLITZ_SESSIONS[u.id] = {
        "active": True,
        "correct": 0,
        "wrong": 0,
        "until": until,
        "job": job_name
    }

    await q.edit_message_text(L["blitz_started"].format(minutes=minutes))
    await send_blitz_poll_app(context.application, q.message.chat_id, uid, u.id, group_id=context.user_data.get("selected_group"))

async def blitz_time_up_job(ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = ctx.job.data["tg_id"]
    sess = BLITZ_SESSIONS.get(tg_id)
    if not sess:
        return
    correct = sess.get("correct", 0)
    wrong = sess.get("wrong", 0)
    total = correct + wrong
    score = correct * POINTS_FOR_CORRECT_BLITZ + wrong * POINTS_FOR_WRONG
    msg = LANGS.get(get_ui_lang(get_or_create_user(tg_id, None)), LANGS["UZ"])["blitz_time_up"].format(
        correct=correct, wrong=wrong, total=total, score=score
    )
    await ctx.bot.send_message(chat_id=tg_id, text=msg)
    BLITZ_SESSIONS.pop(tg_id, None)

async def send_blitz_poll_app(app: Application, chat_id: int, db_user_id: int, tg_user_id: int, group_id: Optional[int] = None):
    word = pick_user_word(db_user_id, group_id)
    if not word:
        await app.bot.send_message(chat_id, t_for(db_user_id, "quiz_no_words"))
        return
    
    # Track this word to avoid consecutive repeats
    LAST_ASKED[(db_user_id, group_id)] = word["id"]
    
    distractors = get_distractors(db_user_id, word["id"], 3, group_id=group_id)
    options = [word["uzbek"]] + distractors
    random.shuffle(options)
    correct_idx = options.index(word["uzbek"])
    q = f"Tarjimasini toping: {word['english']}"
    msg = await app.bot.send_poll(
        chat_id=chat_id,
        question=q,
        options=options,
        type="quiz",
        correct_option_id=correct_idx,
        is_anonymous=False,
        explanation=f"To‘g‘ri javob: {word['uzbek']}",
    )
    if msg.poll is None:
        return
    ACTIVE_POLLS[msg.poll.id] = {
        "chat_id": chat_id,
        "message_id": msg.message_id,
        "tg_user_id": tg_user_id,
        "db_user_id": db_user_id,
        "word_id": word["id"],
        "correct_idx": correct_idx,
        "group_id": group_id,
        "question_num": BLITZ_SESSIONS[tg_user_id].get("correct", 0) + BLITZ_SESSIONS[tg_user_id].get("wrong", 0) + 1,
        "is_blitz": True,
    }

# =====================
# Start / Language / Groups / Dispatch
# =====================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    if is_banned(u.id):
        return


    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    await update.message.reply_text(L["start"], reply_markup=build_main_keyboard(uid))
    await update.message.reply_text(L["choose_lang"], reply_markup=language_keyboard())

async def set_language_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    lang = q.data.split(":")[1]
    uid = get_or_create_user(q.from_user.id, q.from_user.username)
    set_ui_lang(uid, lang)
    kb = build_main_keyboard(uid)
    L = LANGS.get(lang, LANGS["UZ"])
    await q.edit_message_text(L["choose_lang"], reply_markup=None)
    await context.bot.send_message(chat_id=q.message.chat_id, text=L["language_changed_success"], reply_markup=kb)

async def create_group_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    context.user_data["pending_group_create"] = True
    await update.message.reply_text(L["create_group_prompt"])

async def rename_group_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    groups = get_user_groups(uid)
    if not groups:
        await update.message.reply_text(L["no_groups"])
        return
    buttons = [[InlineKeyboardButton(g["name"], callback_data=f"group_rename_select:{g['id']}")] for g in groups]
    await update.message.reply_text(L["group_rename_prompt"], reply_markup=InlineKeyboardMarkup(buttons))

async def delete_group_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    groups = get_user_groups(uid)
    if not groups:
        await update.message.reply_text(L["no_groups"])
        return
    buttons = [[InlineKeyboardButton(g["name"], callback_data=f"group_delete_select:{g['id']}")] for g in groups]
    await update.message.reply_text(L["choose_group_delete"], reply_markup=InlineKeyboardMarkup(buttons))

async def group_io_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    groups = get_user_groups(uid)
    buttons = [[InlineKeyboardButton(g["name"], callback_data=f"group_io_select:{g['id']}:menu")] for g in groups]
    buttons.append([InlineKeyboardButton(L["personal_selected"], callback_data=f"io:{u.id}:menu:none")])
    await update.message.reply_text(L["choose_group_io"], reply_markup=InlineKeyboardMarkup(buttons))

async def select_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    uid = get_or_create_user(u.id, u.username)
    groups = get_user_groups(uid)
    buttons = [
        [InlineKeyboardButton(g["name"], callback_data=f"group_select:{g['id']}")]
        for g in groups
    ]
    buttons.append([InlineKeyboardButton("Personal", callback_data="group_select:personal")])
    prompt = LANGS.get(get_ui_lang(uid), LANGS["UZ"])["group_select_prompt"] if context.user_data.get("pending_action") in ("quiz", "blitz", "add") else LANGS.get(get_ui_lang(uid), LANGS["UZ"])["group_view_prompt"]
    await update.message.reply_text(prompt, reply_markup=InlineKeyboardMarkup(buttons))

async def group_select_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    data = q.data.split(":")[1]
    
    if data == "personal":
        context.user_data["selected_group"] = None
        await q.edit_message_text("All (personal) selected.")
    else:
        group_id = int(data)
        context.user_data["selected_group"] = group_id
        with db() as conn:
            group = conn.execute("SELECT name FROM groups WHERE id=?", (group_id,)).fetchone()
        await q.edit_message_text(f"Group selected: {group['name']}")
    
    pending = context.user_data.get("pending_action")
    if pending == "quiz":
        del context.user_data["pending_action"]
        sess = QUIZ_SESSIONS.get(u.id)
        if sess and sess.get("current_question_num", 1) > 1:
            await context.bot.send_message(q.message.chat_id, t_for(uid, "quiz_continue_prompt"), reply_markup=quiz_continue_kb(get_ui_lang(uid)))
        else:
            await send_quiz_poll(context, q.message.chat_id, uid, u.id, context.user_data.get("selected_group"), question_num=1)
    elif pending == "blitz":
        del context.user_data["pending_action"]
        L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
        await context.bot.send_message(q.message.chat_id, L["blitz_choose_duration"], reply_markup=blitz_duration_kb(get_ui_lang(uid)))
    elif pending == "words":
        del context.user_data["pending_action"]
        await send_words_page(q, uid, u.id, offset=0, group_id=context.user_data.get("selected_group"))
    elif pending == "add":
        del context.user_data["pending_action"]
        await q.edit_message_text(t_for(uid, "add_prompt"))
        context.user_data["awaiting_add"] = context.user_data.get("selected_group")
    else:
        kb = build_main_keyboard(uid)
        await context.bot.send_message(chat_id=q.message.chat_id, text="OK", reply_markup=kb)

async def group_add_select_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    group_id = int(q.data.split(":")[1])
    if not (is_group_owner(uid, group_id) or is_admin(uid)):
        await q.edit_message_text("No permission.")
        return
    context.user_data["awaiting_group_add"] = group_id
    await q.edit_message_text(L["group_add_prompt"])


async def group_add_user_select_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle 'Add user to group' button from groups menu."""
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    group_id = int(q.data.split(":")[1])
    if not (is_group_owner(uid, group_id) or is_admin(uid)):
        await q.edit_message_text("No permission.")
        return
    context.user_data["awaiting_add_user_to_group_from_cb"] = group_id
    await q.edit_message_text(L["add_user_to_group_prompt"])

    
async def create_group_inline_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    context.user_data["pending_group_create"] = True
    try:
        await q.edit_message_text(L["create_group_prompt"])
    except Exception:
        await context.bot.send_message(chat_id=q.message.chat_id, text=L["create_group_prompt"])
    return  

# =====================
# Inline callbacks: stats, words, io, reminders, admin, language, leader, quiz continue
# =====================

def month_keyboard(months: list[str]) -> InlineKeyboardMarkup:
    btns = [[InlineKeyboardButton(m, callback_data=f"stats_month:{m}")] for m in months]
    return InlineKeyboardMarkup(btns or [[InlineKeyboardButton("No data", callback_data="noop")]])

def days_keyboard(days: list[str]) -> InlineKeyboardMarkup:
    btns = [[InlineKeyboardButton(d, callback_data=f"stats_day:{d}")] for d in days]
    btns.append([InlineKeyboardButton("⬅️ Months", callback_data="stats_back_months")])
    return InlineKeyboardMarkup(btns)

async def stats_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)

    data = q.data
    if data.startswith("stats_back_months"):
        months = month_list_for_user(uid)
        await q.edit_message_text("Choose month (YYYY-MM):", reply_markup=month_keyboard(months))
        return

    if data.startswith("stats_month:"):
        ym = data.split(":",1)[1]
        days = days_in_month_for_user(uid, ym)
        if not days:
            await q.edit_message_text(f"No data in month {ym}.")
            return
        await q.edit_message_text(f"Choose day in {ym}:", reply_markup=days_keyboard(days))
        return

    if data.startswith("stats_day:"):
        d = data.split(":",1)[1]
        counts = day_counts(uid, d)
        added = added_words_on(uid, d, limit=20)
        body = (
            f"📊 {d}\n"
            f"✅ Correct: {counts['correct']}\n"
            f"❌ Wrong: {counts['wrong']}\n"
            f"➕ Added words: {counts['added']}\n"
        )
        if added:
            body += "\nAdded (some):\n" + "\n".join(f"• {e} — {u}" for e, u in added)
        await q.edit_message_text(body, reply_markup=days_keyboard(days_in_month_for_user(uid, d[:7])))
        return

async def quiz_continue_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await quiz_continue_handler(update, context)

# Words / io callbacks
async def send_words_page(update_or_query, db_user_id: int, tg_id: int, offset: int = 0, days: Optional[int] = None, group_id: Optional[int] = None):
    total = count_user_words(db_user_id, days=days, group_id=group_id)
    text = words_page_text(db_user_id, offset, total, days=days, group_id=group_id)
    kb = words_page_kb(tg_id, offset, total, days, group_id)
    if isinstance(update_or_query, Update) and update_or_query.message:
        await update_or_query.message.reply_text(text, reply_markup=kb)
    else:
        await update_or_query.edit_message_text(text, reply_markup=kb)

def count_user_words(user_id: int, days: Optional[int] = None, group_id: Optional[int] = None) -> int:
    with db() as conn:
        if group_id is not None:
            if days:
                (n,) = conn.execute("SELECT COUNT(*) FROM words WHERE group_id=? AND date(created_at) >= date('now', ?)", (group_id, f"-{days} day")).fetchone()
            else:
                (n,) = conn.execute("SELECT COUNT(*) FROM words WHERE group_id=?", (group_id,)).fetchone()
        else:
            if days:
                (n,) = conn.execute("SELECT COUNT(*) FROM words WHERE user_id=? AND date(created_at) >= date('now', ?)", (user_id, f"-{days} day")).fetchone()
            else:
                (n,) = conn.execute("SELECT COUNT(*) FROM words WHERE user_id=?", (user_id,)).fetchone()
    return int(n)

def fetch_words_page(user_id: int, offset: int, limit: int = 50, days: Optional[int] = None, group_id: Optional[int] = None) -> list[sqlite3.Row]:
    if offset < 0:
        offset = 0
    with db() as conn:
        if group_id is not None:
            if days:
                rows = conn.execute(
                    "SELECT id, english, uzbek, created_at FROM words WHERE group_id=? AND date(created_at) >= date('now', ?) ORDER BY id DESC LIMIT ? OFFSET ?",
                    (group_id, f"-{days} day", limit, offset)
                ).fetchall()
            else:
                rows = conn.execute("SELECT id, english, uzbek, created_at FROM words WHERE group_id=? ORDER BY id DESC LIMIT ? OFFSET ?", (group_id, limit, offset)).fetchall()
        else:
            if days:
                rows = conn.execute(
                    "SELECT id, english, uzbek, created_at FROM words WHERE user_id=? AND date(created_at) >= date('now', ?) ORDER BY id DESC LIMIT ? OFFSET ?",
                    (user_id, f"-{days} day", limit, offset)
                ).fetchall()
            else:
                rows = conn.execute("SELECT id, english, uzbek, created_at FROM words WHERE user_id=? ORDER BY id DESC LIMIT ? OFFSET ?", (user_id, limit, offset)).fetchall()
    return rows

def words_page_text(user_id: int, offset: int, total: int, days: Optional[int] = None, group_id: Optional[int] = None, for_username: Optional[str] = None) -> str:
    rows = fetch_words_page(user_id, offset, days=days, group_id=group_id)
    if not rows:
        return "No words yet."
    start = offset + 1
    end = min(offset + 50, total)
    body = "\n".join(f"• {r['english']} — {r['uzbek']}" for r in rows)
    rng = "All" if not days else (f"Last {days} days")
    header = f"🗒 Words ({start}–{end}/{total}, {rng})\n"
    if for_username:
        header = f"🗒 {for_username}'s words ({start}–{end}/{total}, {rng})\n"
    return header + body

def words_page_kb(tg_id: int, offset: int, total: int, days: Optional[int], group_id: Optional[int]) -> InlineKeyboardMarkup:
    buttons = []
    row = []
    prev_off = max(offset - 50, 0)
    next_off = offset + 50 if offset + 50 < total else None

    if offset > 0:
        row.append(InlineKeyboardButton("⬅️ Previous 50", callback_data=f"w:{tg_id}:{prev_off}:{days or 'all'}:{group_id or 'none'}"))
    row.append(InlineKeyboardButton("🗂 Filter", callback_data=f"wf:{tg_id}:menu:{group_id or 'none'}"))
    row.append(InlineKeyboardButton("🗑 Delete mode", callback_data=f"wd:{tg_id}:{offset}:{days or 'all'}:{group_id or 'none'}"))
    
    row.append(InlineKeyboardButton("📁 Import/Export", callback_data=f"io:{tg_id}:menu:{group_id or 'none'}"))
    row.append(InlineKeyboardButton("🗑 Clear all", callback_data=f"wclear:{tg_id}:{group_id or 'none'}"))
    if next_off is not None:
        row.append(InlineKeyboardButton("Next 50 ➡️", callback_data=f"w:{tg_id}:{next_off}:{days or 'all'}:{group_id or 'none'}"))
    buttons.append(row)
    if days in (7,30):
        buttons.append([InlineKeyboardButton(f"Filter: Last {days} days", callback_data="noop")])
    return InlineKeyboardMarkup(buttons)

# Admin words page
async def send_admin_words_page(query, target_user_id: int, offset: int = 0, days: Optional[int] = None):
    with db() as conn:
        r = conn.execute("SELECT username FROM users WHERE id=?", (target_user_id,)).fetchone()
    uname = r["username"] or "Anon"
    total = count_user_words(target_user_id, days=days)
    text = words_page_text(target_user_id, offset, total, days=days)
    kb = admin_words_page_kb(target_user_id, offset, total, days)
    await query.edit_message_text(text, reply_markup=kb)

def admin_words_page_kb(target_user_id: int, offset: int, total: int, days: Optional[int]) -> InlineKeyboardMarkup:
    buttons = []
    row = []
    prev_off = max(offset - 50, 0)
    next_off = offset + 50 if offset + 50 < total else None

    if offset > 0:
        row.append(InlineKeyboardButton("⬅️ Previous 50", callback_data=f"admin:words:{target_user_id}:{prev_off}:{days or 'all'}"))
    row.append(InlineKeyboardButton("🗂 Filter", callback_data=f"admin:wf:{target_user_id}:menu"))
    row.append(InlineKeyboardButton("🗑 Delete mode", callback_data=f"admin:wd:{target_user_id}:{offset}:{days or 'all'}"))
    if next_off is not None:
        row.append(InlineKeyboardButton("Next 50 ➡️", callback_data=f"admin:words:{target_user_id}:{next_off}:{days or 'all'}"))
    buttons.append(row)
    buttons.append([InlineKeyboardButton("⬅️ User info", callback_data=f"admin:user_info:{target_user_id}")])
    if days in (7,30):
        buttons.append([InlineKeyboardButton(f"Filter: Last {days} days", callback_data="noop")])
    return InlineKeyboardMarkup(buttons)

async def words_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    data = q.data
    parts = data.split(":")
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])

    if data.startswith("wclear:"):
        tg_id_str = parts[1]
        gid_s = parts[2] if len(parts) > 2 else "none"
        if str(u.id) != tg_id_str:
            await q.answer("This list is not yours.", show_alert=True)
            return
        group_id = None if gid_s == "none" else int(gid_s)
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(L["clear_all_yes"], callback_data=f"wclear_confirm:{tg_id_str}:{gid_s}:yes"),
             InlineKeyboardButton(L["clear_all_no"], callback_data=f"wclear_confirm:{tg_id_str}:{gid_s}:no")]
        ])
        await q.edit_message_text(L["clear_all_confirm"], reply_markup=kb)
        return

    if data.startswith("wclear_confirm:"):
        tg_id_str = parts[1]
        gid_s = parts[2] if len(parts) > 2 else "none"
        confirm = parts[3] if len(parts) > 3 else "no"
        if str(u.id) != tg_id_str:
            await q.answer("Not yours.", show_alert=True)
            return
        group_id = None if gid_s == "none" else int(gid_s)
        if confirm == "yes":
            ok = delete_all_words(uid, group_id=group_id)
            msg = L["clear_all_success"] if ok else L["clear_all_fail"]
            await q.edit_message_text(msg)
        else:
            await q.edit_message_text("Cancelled.")
        return

    if data.startswith("wf:"):
        tg_id_str = parts[1]
        gid_s = parts[3] if len(parts) > 3 else "none"
        if str(u.id) != tg_id_str:
            await q.answer("This list is not yours.", show_alert=True)
            return
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("All", callback_data=f"wfr:{tg_id_str}:all:{gid_s}"),
             InlineKeyboardButton("Last 7 days", callback_data=f"wfr:{tg_id_str}:7:{gid_s}"),
             InlineKeyboardButton("Last 30 days", callback_data=f"wfr:{tg_id_str}:30:{gid_s}")],
            [InlineKeyboardButton("⬅️ Back", callback_data=f"w:{tg_id_str}:0:all:{gid_s}")]
        ])
        await q.edit_message_text("Choose filter:", reply_markup=kb)
        return

    if data.startswith("wfr:"):
        tg_id_str = parts[1]
        rng = parts[2]
        gid_s = parts[3] if len(parts) > 3 else "none"
        if str(u.id) != tg_id_str:
            await q.answer("Not yours.", show_alert=True)
            return
        days = None
        if rng == "7": days = 7
        elif rng == "30": days = 30
        group_id = None if gid_s == "none" else int(gid_s)
        context.user_data["words_days"] = days
        await send_words_page(q, uid, u.id, offset=0, days=days, group_id=group_id)
        return

    if data.startswith("wd:"):
        tg_id_str = parts[1]
        offset_s = parts[2]
        days_s = parts[3] if len(parts) > 3 else "all"
        gid_s = parts[4] if len(parts) > 4 else "none"
        if str(u.id) != tg_id_str:
            await q.answer("This list is not yours.", show_alert=True)
            return
        try:
            offset = max(int(offset_s), 0)
        except:
            offset = 0
        days = None if days_s == "all" else int(days_s)
        group_id = None if gid_s == "none" else int(gid_s)
        rows = fetch_words_page(uid, offset, days=days, group_id=group_id)
        if not rows:
            await q.edit_message_text(L["no_words"])
            return
        lines = [f"{i+1}. {r['english']} — {r['uzbek']}" for i,r in enumerate(rows)]
        body = "🗑 " + L["delete_mode"] + "\n" + "\n".join(lines) + "\n\n" + L["awaiting_delete_number"]
        nav = [[InlineKeyboardButton("⬅️ Back", callback_data=f"w:{tg_id_str}:{offset}:{days_s}:{gid_s}"), InlineKeyboardButton("Close", callback_data=f"w:{tg_id_str}:close")]]
        context.user_data["delete_mode_data"] = {
            "tg_id": u.id,
            "offset": offset,
            "days": days,
            "group_id": group_id,
            "rows": rows,
            "uid": uid
        }
        await q.edit_message_text(body, reply_markup=InlineKeyboardMarkup(nav))
        context.user_data["awaiting_delete_number"] = True
        return

    if data.startswith("wdx:"):
        tg_id_str = parts[1]
        offset_s = parts[2]
        wid_s = parts[3]
        days_s = parts[4] if len(parts) > 4 else "all"
        gid_s = parts[5] if len(parts) > 5 else "none"
        if str(u.id) != tg_id_str:
            await q.answer("Not yours.", show_alert=True)
            return
        ok = delete_word_if_owner(int(wid_s), uid)
        msg = "🗑 Deleted." if ok else "Not found or no permission."
        days = None if days_s == "all" else int(days_s)
        group_id = None if gid_s == "none" else int(gid_s)
        await q.answer(msg, show_alert=False)
        await send_words_page(q, uid, int(tg_id_str), offset=int(offset_s), days=days, group_id=group_id)
        return

    if data.startswith("io:"):
        tg_id_str = parts[1]
        action = parts[2]
        gid_s = parts[3] if len(parts) > 3 else "none"
        if str(u.id) != tg_id_str:
            await q.answer("Not yours.", show_alert=True)
            return
        group_id = None if gid_s == "none" else int(gid_s)
        if action == "menu":
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("📤 Export XLSX", callback_data=f"io:{tg_id_str}:export:{gid_s}"),
                 InlineKeyboardButton("📥 Import (upload XLSX)", callback_data=f"io:{tg_id_str}:import:{gid_s}")],
                [InlineKeyboardButton("⬅️ Back", callback_data=f"w:{tg_id_str}:0:all:{gid_s}")]
            ])
            await q.edit_message_text(L["import_export_text"], reply_markup=kb)
            return
        if action == "export":
            with db() as conn:
                if group_id is not None:
                    rows = conn.execute("SELECT english, uzbek FROM words WHERE group_id=? ORDER BY id DESC", (group_id,)).fetchall()
                else:
                    rows = conn.execute("SELECT english, uzbek FROM words WHERE user_id=? ORDER BY id DESC", (uid,)).fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["English", "Uzbek"])
            for r in rows:
                ws.append([r["english"], r["uzbek"]])
            path = f"words_{u.id}_{gid_s}.xlsx"
            wb.save(path)
            await q.message.chat.send_document(document=open(path,"rb"), filename=path, caption="XLSX ready (A=English, B=Uzbek).")
            os.remove(path)
            return
        if action == "import":
            context.user_data["awaiting_import"] = {"group_id": group_id}
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="import_cancel")]])
            await q.edit_message_text(L["import_prompt"], reply_markup=kb)
            return

    if data.startswith("group_io_select:"):
        group_id = int(parts[1])
        action = parts[2]
        if not is_group_owner(uid, group_id):
            await q.answer("You are not the owner of this group.", show_alert=True)
            return
        if action == "menu":
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("📤 Export XLSX", callback_data=f"group_io_select:{group_id}:export"),
                 InlineKeyboardButton("📥 Import (upload XLSX)", callback_data=f"group_io_select:{group_id}:import")],
                [InlineKeyboardButton("☁️ Cloud Backup", callback_data=f"group_io_select:{group_id}:cloud_backup")],
                [InlineKeyboardButton("⬅️ Back", callback_data="group_io_back")]
            ])
            await q.edit_message_text(L["import_export_text"], reply_markup=kb)
        elif action == "export":
            with db() as conn:
                rows = conn.execute("SELECT english, uzbek FROM words WHERE group_id=? ORDER BY id DESC", (group_id,)).fetchall()
            if not rows:
                await q.edit_message_text("No words in the group.")
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["English", "Uzbek"])
            for r in rows:
                ws.append([r["english"], r["uzbek"]])
            path = f"words_group_{group_id}.xlsx"
            wb.save(path)
            await q.message.chat.send_document(document=open(path, "rb"), filename=path, caption="Group words XLSX (A=English, B=Uzbek).")
            os.remove(path)
        elif action == "import":
            context.user_data["awaiting_group_import"] = group_id
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="group_io_cancel")]])
            await q.edit_message_text(L["import_prompt"], reply_markup=kb)
        elif action == "cloud_backup":
            # User-facing cloud backup feature: creates personal backup of user's data
            try:
                backup_file = create_user_data_backup(uid)
                file_size = os.path.getsize(backup_file) / (1024 * 1024)  # Convert to MB
                caption = f"☁️ Your personal data backup\n📦 Size: {file_size:.2f} MB\n⏰ Created: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                await q.message.chat.send_document(
                    document=open(backup_file, "rb"),
                    filename=f"wordl_backup_{uid}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    caption=caption
                )
                await q.answer("✅ Cloud backup created and sent!", show_alert=False)
                log.info(f"User {uid} created personal backup: {backup_file}")
            except Exception as e:
                log.error(f"Cloud backup failed for user {uid}: {e}")
                await q.answer(f"❌ Backup failed: {str(e)}", show_alert=True)
        return

    if data == "group_io_back":
        groups = get_user_groups(uid)
        buttons = [
            [InlineKeyboardButton(g["name"], callback_data=f"group_io_select:{g['id']}:menu")]
            for g in groups
        ]
        buttons.append([InlineKeyboardButton("Personal words", callback_data=f"io:{u.id}:menu")])
        await q.edit_message_text("Choose group for Import/Export or personal words:", reply_markup=InlineKeyboardMarkup(buttons))
        return

    if data == "group_io_cancel":
        context.user_data["awaiting_group_import"] = None
        await q.edit_message_text("Group import cancelled.")
        return

    if data.startswith("w:"):
        tg_id_str = parts[1]
        action = parts[2]
        days_s = parts[3] if len(parts) > 3 else "all"
        gid_s = parts[4] if len(parts) > 4 else "none"
        if str(u.id) != tg_id_str:
            await q.answer("This list is not yours.", show_alert=True)
            return
        if action == "close":
            try: await q.message.delete()
            except Exception: pass
            return
        try:
            offset = max(int(action), 0)
        except ValueError:
            offset = 0
        days = None if days_s == "all" else int(days_s)
        group_id = None if gid_s == "none" else int(gid_s)
        total = count_user_words(uid, days=days, group_id=group_id)
        if offset >= total:
            offset = max(total - (total % 50 or 50), 0)
        await send_words_page(q, uid, int(tg_id_str), offset=offset, days=days, group_id=group_id)
        return

async def import_cancel_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data["awaiting_import"] = None
    await q.edit_message_text(t_for(get_or_create_user(q.from_user.id, q.from_user.username), "import_cancelled"))

async def admin_words_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    data = q.data
    parts = data.split(":")
    admin_u = q.from_user
    if not is_admin(admin_u.id):
        await q.answer("For admins only", show_alert=True)
        return
    target_uid = int(parts[2])
    action = parts[3]
    days_s = parts[4] if len(parts) > 4 else "all"

    if data.startswith("admin:wf:"):
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("All", callback_data=f"admin:wfr:{target_uid}:all"),
             InlineKeyboardButton("Last 7 days", callback_data=f"admin:wfr:{target_uid}:7"),
             InlineKeyboardButton("Last 30 days", callback_data=f"admin:wfr:{target_uid}:30")],
            [InlineKeyboardButton("⬅️ Back", callback_data=f"admin:words:{target_uid}:0:all")]
        ])
        await q.edit_message_text("Choose filter:", reply_markup=kb)
        return

    if data.startswith("admin:wfr:"):
        rng = parts[3]
        days = None
        if rng == "7": days = 7
        elif rng == "30": days = 30
        await send_admin_words_page(q, target_uid, offset=0, days=days)
        return

    if data.startswith("admin:wd:"):
        offset_s = parts[3]
        days_s = parts[4] if len(parts) > 4 else "all"
        try:
            offset = max(int(offset_s), 0)
        except:
            offset = 0
        days = None if days_s == "all" else int(days_s)
        rows = fetch_words_page(target_uid, offset, days=days)
        if not rows:
            await q.edit_message_text("No words found.")
            return
        lines = [f"{i+1}. {r['english']} — {r['uzbek']} [#{r['id']}]" for i,r in enumerate(rows)]
        body = "🗑 Delete mode (this page):\n" + "\n".join(lines)
        btns = []
        for r in rows[:20]:
            btns.append([InlineKeyboardButton(f"❌ #{r['id']}", callback_data=f"admin:wdx:{target_uid}:{offset}:{r['id']}:{days_s}")])
        nav = [InlineKeyboardButton("⬅️ Back", callback_data=f"admin:words:{target_uid}:{offset}:{days_s}"), InlineKeyboardButton("Close", callback_data="admin:close")]
        btns.append(nav)
        await q.edit_message_text(body, reply_markup=InlineKeyboardMarkup(btns))
        return

    if data.startswith("admin:wdx:"):
        offset_s = parts[3]
        wid_s = parts[4]
        days_s = parts[5] if len(parts) > 5 else "all"
        ok = admin_delete_word(int(wid_s))
        msg = "🗑 Deleted." if ok else "Not found."
        days = None if days_s == "all" else int(days_s)
        await q.answer(msg, show_alert=False)
        await send_admin_words_page(q, target_uid, offset=int(offset_s), days=days)
        return

    try:
        offset = max(int(action), 0)
    except ValueError:
        offset = 0
    days = None if days_s == "all" else int(days_s)
    total = count_user_words(target_uid, days=days)
    if offset >= total:
        offset = max(total - (total % 50 or 50), 0)
    await send_admin_words_page(q, target_uid, offset=offset, days=days)

# =====================
# Import document handler
# =====================

async def on_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    XLSX import handler — faqat .xlsx fayllar qabul qiladi.
    Kutish: context.user_data may contain:
      - "awaiting_import" (bool or dict, optionally with "group_id")
      - "awaiting_group_import" (group_id)
    """
    import os
    import tempfile

    u = update.effective_user
    if is_banned(u.id):
        return
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])

    awaiting = context.user_data.get("awaiting_import")
    awaiting_group = context.user_data.get("awaiting_group_import")

    # Agar import rejimi yo'q bo'lsa — hech narsa qilinmasin
    if not (awaiting or awaiting_group):
        return

    doc = update.message.document
    if not doc:
        await update.message.reply_text("Hech qanday hujjat topilmadi.")
        return

    fname = (doc.file_name or "").lower()
    if not fname.endswith(".xlsx"):
        await update.message.reply_text("Iltimos, faqat .xlsx formatidagi fayl yuboring (English, Uzbek).")
        return

    # Fayl hajmi cheklovi (xavfsizlik uchun)
    MAX_MB = 20
    if doc.file_size and doc.file_size > MAX_MB * 1024 * 1024:
        await update.message.reply_text(f"Fayl hajmi {MAX_MB}MB dan katta — import bekor qilindi.")
        return

    # Yaqin vaqt ichida yuklab olish va ishlov berish
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)

    try:
        # Yuklash (network xatolari uchun try/except)
        try:
            file = await doc.get_file()
            await file.download_to_drive(tmp_path)
        except Exception as e:
            # Log qilish (agar log obyekti mavjud bo'lsa)
            try:
                log.exception("Failed to download document")
            except Exception:
                pass
            await update.message.reply_text(f"Faylni olishda xatolik: {e}")
            return

        # XLSXni ochish
        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        except Exception as e:
            try:
                log.exception("Failed to open workbook")
            except Exception:
                pass
            await update.message.reply_text(f"XLSX ochilmadi: {e}")
            return

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # Agar jadval boshida sarlavha bo'lsa (English / Uzbek) — olib tashlash
        if rows and len(rows) >= 1:
            c0 = rows[0][0] if rows[0] and len(rows[0]) > 0 else None
            c1 = rows[0][1] if rows[0] and len(rows[0]) > 1 else None
            c0s = str(c0).strip().lower() if c0 is not None else ""
            c1s = str(c1).strip().lower() if c1 is not None else ""
            if ("english" in c0s) and ("uzb" in c1s or "uzbek" in c1s):
                rows = rows[1:]

        inserted = 0
        errors = []

        # Maqsad guruhni aniqlash (agar mavjud bo'lsa)
        target_group = None
        if awaiting_group:
            target_group = awaiting_group
            # guruh egasini tekshirish
            if not is_group_owner(uid, target_group):
                await update.message.reply_text("Siz ushbu guruh egasi emassiz; import bekor qilindi.")
                return
            context.user_data["awaiting_group_import"] = None
        else:
            # awaiting shartida dict bo'lishi mumkin
            if isinstance(awaiting, dict) and awaiting.get("group_id"):
                target_group = awaiting.get("group_id")
            context.user_data["awaiting_import"] = False

        # Har bir qatorni qayta ishlash (batch processing to avoid timeouts)
        batch_size = 100
        for batch_start in range(0, len(rows), batch_size):
            batch_end = min(batch_start + batch_size, len(rows))
            for idx in range(batch_start, batch_end):
                r = rows[idx]
                row_num = idx + 1
                try:
                    if not r or len(r) < 2:
                        errors.append(f"Row {row_num}: enough columns yo'q.")
                        continue
                    raw_eng = r[0]
                    raw_uz = r[1]
                    eng = "" if raw_eng is None else str(raw_eng).strip()
                    uz = "" if raw_uz is None else str(raw_uz).strip()
                    if not eng or not uz:
                        errors.append(f"Row {row_num}: bo'sh ENG yoki UZ.")
                        continue
                    # Inglizcha so'z (butun ibora) olinadi
                    eng_word = eng
                    # So'zni DBga qo'shish (mavjud add_word funksiyasini ishlatadi)
                    add_word(uid, eng_word, uz, group_id=target_group)
                    inserted += 1
                except Exception as e:
                    try:
                        log.exception("Error adding word from xlsx")
                    except Exception:
                        pass
                    errors.append(f"Row {row_num}: {e}")
            # Small delay between batches to prevent timeouts
            if batch_end < len(rows):
                await asyncio.sleep(0.1)

        # Javob xabarini tayyorlash
        import_done_msg = L.get("import_done", "{n} words imported").format(n=inserted)
        if errors:
            preview = "\n".join(errors[:10])
            if len(errors) > 10:
                preview += f"\n... +{len(errors)-10} boshqa xatoliklar."
            import_done_msg += f"\n\nBa'zi qatorlar import qilinmadi:\n{preview}"

        await update.message.reply_text(import_done_msg)
    finally:
        # vaqtincha faylni tozalash
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


# =====================
# Reminders UI callbacks (reminder_cb)
# =====================

async def open_reminder_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    uid = get_or_create_user(u.id, u.username)
    st = get_settings(uid)
    text = LANGS.get(st.get("ui_lang", "UZ"))["reminder_panel"].format(
        state=("Enabled" if st["remind_enabled"] else "Disabled"),
        goal=st["daily_goal"], time=st["remind_time"]
    )
    await update.message.reply_text(text, reply_markup=reminder_kb(bool(st['remind_enabled']), st['daily_goal'], st['remind_time']))

async def reminder_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    st = get_settings(uid)
    data = q.data
    if data == "rem:close":
        try: await q.message.delete()
        except Exception: pass
        return
    if data == "rem:toggle":
        new_state = 0 if st["remind_enabled"] else 1
        set_settings(uid, remind_enabled=new_state)
        st["remind_enabled"] = new_state
        schedule_user_reminder(context.application, u.id, st["remind_time"], bool(new_state))
    elif data.startswith("rem:goal:"):
        goal_action = data.split(":",2)[2]
        if goal_action == "custom":
            context.user_data["awaiting_custom_goal"] = True
            L = LANGS.get(st.get("ui_lang", "UZ"), LANGS["UZ"])
            await q.edit_message_text(L["send_goal_prompt"])
            return
    elif data.startswith("rem:time:"):
        hhmm = data.split(":",2)[2]
        if hhmm == "custom":
            context.user_data["awaiting_custom_time"] = True
            L = LANGS.get(st.get("ui_lang", "UZ"), LANGS["UZ"])
            await q.edit_message_text(L["send_time_prompt"])
            return
        set_settings(uid, remind_time=hhmm)
        st["remind_time"] = hhmm
        if st["remind_enabled"]:
            schedule_user_reminder(context.application, u.id, hhmm, True)
    text = LANGS.get(st.get("ui_lang","UZ"))["reminder_panel"].format(
        state=("Enabled" if st["remind_enabled"] else "Disabled"),
        goal=st["daily_goal"], time=st["remind_time"]
    )
    try:
        await q.edit_message_text(text, reply_markup=reminder_kb(bool(st['remind_enabled']), st['daily_goal'], st['remind_time']))
    except BadRequest as e:
        if "not modified" in str(e).lower():
            pass
        else:
            raise


# =====================
# Settings UI
# =====================
def settings_kb(st: dict) -> InlineKeyboardMarkup:
    # Build keyboard for settings: profile, quiz repeat, restart on incorrect
    kb = []
    kb.append([InlineKeyboardButton("👤 Profile", callback_data="settings:profile")])
    # quick quiz_repeat options
    kb.append([
        InlineKeyboardButton("1", callback_data="settings:quiz_repeat:1"),
        InlineKeyboardButton("2", callback_data="settings:quiz_repeat:2"),
        InlineKeyboardButton("3", callback_data="settings:quiz_repeat:3"),
        InlineKeyboardButton("5", callback_data="settings:quiz_repeat:5")
    ])
    kb.append([InlineKeyboardButton("Custom", callback_data="settings:quiz_repeat:custom")])
    # restart on incorrect options
    kb.append([
        InlineKeyboardButton("Restart:1", callback_data="settings:restart:1"),
        InlineKeyboardButton("Restart:2", callback_data="settings:restart:2"),
        InlineKeyboardButton("Restart:3", callback_data="settings:restart:3")
    ])
    kb.append([InlineKeyboardButton("Custom", callback_data="settings:restart:custom")])
    kb.append([InlineKeyboardButton("⬅️ Close", callback_data="settings:close")])
    return InlineKeyboardMarkup(kb)


async def open_settings_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    uid = get_or_create_user(u.id, u.username)
    st = get_settings(uid)
    L = LANGS.get(st.get("ui_lang", "UZ"), LANGS["UZ"])
    profile = L.get("profile_info", "").format(username=u.username or "", tg_id=u.id, uid=uid)
    text = L.get("settings_panel").format(profile=profile, quiz_repeat=st.get("quiz_repeat",1), restart_on_incorrect=st.get("restart_on_incorrect",3))
    await update.message.reply_text(text, reply_markup=settings_kb(st))


async def settings_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    st = get_settings(uid)
    L = LANGS.get(st.get("ui_lang", "UZ"), LANGS["UZ"])
    data = q.data
    if data == "settings:close":
        try: await q.message.delete()
        except Exception: pass
        return
    if data == "settings:profile":
        profile = L.get("profile_info").format(username=u.username or "", tg_id=u.id, uid=uid)
        try:
            await q.edit_message_text(profile, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="settings:back")]]))
        except BadRequest:
            pass
        return
    if data == "settings:back":
        st = get_settings(uid)
        profile = L.get("profile_info").format(username=u.username or "", tg_id=u.id, uid=uid)
        text = L.get("settings_panel").format(profile=profile, quiz_repeat=st.get("quiz_repeat",1), restart_on_incorrect=st.get("restart_on_incorrect",3))
        try:
            await q.edit_message_text(text, reply_markup=settings_kb(st))
        except BadRequest:
            pass
        return
    # quiz_repeat actions
    if data.startswith("settings:quiz_repeat:"):
        val = data.split(":",2)[2]
        if val == "custom":
            context.user_data["awaiting_quiz_repeat"] = True
            await q.edit_message_text(L["send_quiz_repeat_prompt"])
            return
        try:
            v = int(val)
            set_settings(uid, quiz_repeat=v)
            st["quiz_repeat"] = v
            await q.edit_message_text(L["settings_changed"], reply_markup=settings_kb(st))
        except Exception:
            await q.answer(L["invalid_number"], show_alert=True)
        return
    # restart actions
    if data.startswith("settings:restart:"):
        val = data.split(":",2)[2]
        if val == "custom":
            context.user_data["awaiting_restart_incorrect"] = True
            await q.edit_message_text(L["send_restart_prompt"])
            return
        try:
            v = int(val)
            set_settings(uid, restart_on_incorrect=v)
            st["restart_on_incorrect"] = v
            await q.edit_message_text(L["settings_changed"], reply_markup=settings_kb(st))
        except Exception:
            await q.answer(L["invalid_number"], show_alert=True)
        return

# =====================
# Admin callbacks
# =====================

def admin_menu_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔊 Broadcast", callback_data="admin:bc"),
         InlineKeyboardButton("👥 Users", callback_data="admin:users:0:active:id")],
        [InlineKeyboardButton("📈 Global stats", callback_data="admin:stats"),
         InlineKeyboardButton("📦 XLSX export (all)", callback_data="admin:export")],
        [InlineKeyboardButton("💾 Backup", callback_data="admin:backup"),
         InlineKeyboardButton("🔄 Restore", callback_data="admin:restore")],
        [InlineKeyboardButton("🗑️ Delete Backups", callback_data="admin:delete_backups")],
        [InlineKeyboardButton("👥 Groups", callback_data="admin:groups")],
        [InlineKeyboardButton("⬅️ Close", callback_data="admin:close")]
    ])

def admin_groups_submenu_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👥 Groups list", callback_data="admin:groups:0")],
        [InlineKeyboardButton("🛠 Manage groups", callback_data="admin:manage_groups")],
        [InlineKeyboardButton("➕ Create group", callback_data="admin:create_group")],
        [InlineKeyboardButton("➕ Add word to group", callback_data="admin:group_add_word")],
        [InlineKeyboardButton("⬅️ Admin menu", callback_data="admin:main")]
    ])

async def open_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message:
        await update.message.reply_text("Admin panel:", reply_markup=admin_menu_kb())

async def admin_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    uid = q.from_user.id

    if not is_admin(uid):
        await q.answer("For admins only", show_alert=True)
        return

    data = q.data or ""

    if data == "admin:close":
        try:
            await q.message.delete()
        except Exception:
            pass
        return

    if data == "admin:main":
        L = LANGS.get(get_ui_lang(get_or_create_user(q.from_user.id, q.from_user.username)), LANGS["UZ"])
        await q.edit_message_text(L["admin_panel_heading"], reply_markup=admin_menu_kb())
        return

    if data == "admin:bc":
        context.user_data["admin_mode"] = "bc_wait_text"
        await q.edit_message_text(
            "🔊 Send broadcast text (text only).",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")]])
        )
        return

    if data == "admin:cancel":
        context.user_data["admin_mode"] = None
        context.user_data["bc_text"] = None
        context.user_data["admin_await_user"] = False
        await q.edit_message_text("Cancelled.", reply_markup=admin_menu_kb())
        return

    if data == "admin:bc_send":
        txt = context.user_data.get("bc_text")
        context.user_data["admin_mode"] = None
        context.user_data["bc_text"] = None
        if not txt:
            await q.edit_message_text("Text not found.", reply_markup=admin_menu_kb())
            return

        ids = iter_all_tg_ids()
        ok = 0
        fail = 0
        for i, tid in enumerate(ids, start=1):
            try:
                await context.bot.send_message(tid, txt)
                ok += 1
            except Exception:
                fail += 1
            if i % 25 == 0:
                await asyncio.sleep(1.0)

        await q.edit_message_text(f"Sent ✅: {ok}\nError ❌: {fail}", reply_markup=admin_menu_kb())
        return

    # === USERS LIST ===
    if data.startswith("admin:users:"):
        parts = data.split(":")
        try:
            offset = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
        except Exception:
            offset = 0

        active_only = (parts[3] == "active") if len(parts) > 3 else True
        sort_by = parts[4] if len(parts) > 4 else "id"

        total = count_users(active_only=active_only)
        rows = fetch_users_page(offset, active_only=active_only, sort_by=sort_by)

        text = f"👥 <b>Foydalanuvchilar roʻyxati</b>\n📄 Jami: <b>{total}</b>\n\n"
        kb_rows = []

        for r in rows:
            user_db_id = r["id"]
            tg_id = r["tg_id"]
            username = r["username"]
            first_seen = (r["first_seen"] or "")[:10]

            if username:
                user_link = f'<a href="tg://user?id={tg_id}">@{html.escape(username)}</a>'
                display_name = f"@{username}"
            else:
                user_link = f'<a href="tg://user?id={tg_id}">User {tg_id}</a>'
                display_name = "Nick"

            text += f"{user_db_id}. {user_link} — {first_seen}\n"

            kb_rows.append([
                InlineKeyboardButton(f"⚙️ Sozlamalar — {display_name}", callback_data=f"admin:user_info:{user_db_id}")
            ])

        step = 20
        nav = []
        if offset > 0:
            nav.append(InlineKeyboardButton("⬅️ Prev", callback_data=f"admin:users:{max(offset-step,0)}:{'active' if active_only else 'all'}:{sort_by}"))
        if offset + step < total:
            nav.append(InlineKeyboardButton("Next ➡️", callback_data=f"admin:users:{offset+step}:{'active' if active_only else 'all'}:{sort_by}"))
        if nav:
            kb_rows.append(nav)

        kb_rows.append([InlineKeyboardButton("🔙 Back", callback_data="admin:main")])

        try:
            await q.edit_message_text(
                text,
                reply_markup=InlineKeyboardMarkup(kb_rows),
                parse_mode="HTML",
                disable_web_page_preview=True
            )
        except BadRequest as e:
            msg = str(e).lower()
            if "modified" in msg or "entities" in msg:
                pass
            else:
                raise
        return

    # === SINGLE USER ACTIONS ===
    if data.startswith("admin:user:"):
        parts = data.split(":")
        if len(parts) < 4:
            await q.edit_message_text("Invalid format.", reply_markup=admin_menu_kb())
            return

        try:
            user_db_id = int(parts[2])
        except Exception:
            await q.edit_message_text("ID invalid.", reply_markup=admin_menu_kb())
            return

        action = parts[3]

        with db() as conn:
            row = conn.execute("SELECT * FROM users WHERE id=?", (user_db_id,)).fetchone()
        if not row:
            await q.edit_message_text("Not found.", reply_markup=admin_menu_kb())
            return

        tg_id = row["tg_id"]

        try:
            if action == "toggle_active":
                new_active = 0 if row["active"] else 1
                set_user_active(tg_id, bool(new_active))
            elif action == "toggle_admin":
                new_role = "admin" if row["role"] == "user" else "user"
                set_user_role(tg_id, new_role)
            elif action == "add_word":
                context.user_data["admin_mode"] = "user_add_word"
                await q.edit_message_text(
                    "Send user ID and word: user_id:english - translation\nExample: 1:Hello - Hello",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")]])
                )
                return
            elif action == "edit_points":
                context.user_data["admin_mode"] = "user_edit_points"
                await q.edit_message_text(
                    "Send user ID and new points: user_id:points\nExample: 1:100",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")]])
                )
                return
        except Exception as e:
            await q.edit_message_text(f"Error during operation: {e}", reply_markup=admin_menu_kb())
            return

        text = get_user_info_text(user_db_id)
        kb = get_user_info_kb(user_db_id)
        await q.edit_message_text(text, reply_markup=kb)
        return

    # === USER SEARCH ===
    if data == "admin:user_search":
        context.user_data["admin_await_user"] = True
        await q.edit_message_text(
            "Send user TG ID or username to search:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")]])
        )
        return

    # === USER INFO (DETAIL VIEW) ===
    if data.startswith("admin:user_info:"):
        parts = data.split(":")
        if len(parts) < 3:
            await q.edit_message_text("ID not found.", reply_markup=admin_menu_kb())
            return
        try:
            user_db_id = int(parts[2])
        except Exception:
            await q.edit_message_text("ID invalid.", reply_markup=admin_menu_kb())
            return

        text = get_user_info_text(user_db_id)
        kb = get_user_info_kb(user_db_id)
        await q.edit_message_text(text, reply_markup=kb)
        return

    # === STATISTICS ===
    if data == "admin:stats":
        text = (
            f"👥 Users: {count_users()}\n"
            f"🗒 Words: {count_words_all()}\n"
            f"➕ Added: {count_stats_all('added')}\n"
            f"✅ Correct: {count_stats_all('correct')}\n"
            f"❌ Wrong: {count_stats_all('wrong')}"
        )
        await q.edit_message_text(text, reply_markup=admin_menu_kb())
        return

    # === EXPORT DATA ===
    if data == "admin:export":
        users_path = f"export_users_{uid}.xlsx"
        words_path = f"export_words_{uid}.xlsx"
        stats_path = f"export_stats_{uid}.xlsx"
        try:
            with db() as conn:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["id","tg_id","username","first_seen","active","role","points"])
                for r in conn.execute("SELECT id,tg_id,username,first_seen,active,role,points FROM users ORDER BY id"):
                    ws.append([r["id"], r["tg_id"], r["username"] or "", r["first_seen"], r["active"], r["role"], r["points"]])
                wb.save(users_path)

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["id","user_id","group_id","english","uzbek","created_at","review_level","next_review"])
                for r in conn.execute("SELECT id,user_id,group_id,english,uzbek,created_at,review_level,next_review FROM words ORDER BY id"):
                    ws.append([r["id"], r["user_id"] or "", r["group_id"] or "", r["english"], r["uzbek"], r["created_at"], r["review_level"], r["next_review"]])
                wb.save(words_path)

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["id","user_id","action","word_id","created_at","local_date"])
                for r in conn.execute("SELECT id,user_id,action,word_id,created_at,local_date FROM stats ORDER BY id"):
                    ws.append([r["id"], r["user_id"], r["action"], r["word_id"] or "", r["created_at"], r["local_date"]])
                wb.save(stats_path)

            await q.edit_message_text("📦 XLSX files ready. Sending...", reply_markup=admin_menu_kb())

            chat_id = q.message.chat.id
            for path in (users_path, words_path, stats_path):
                with open(path, "rb") as f:
                    await context.bot.send_document(chat_id=chat_id, document=f, filename=os.path.basename(path))
        except Exception as e:
            await q.edit_message_text(f"Error: {e}", reply_markup=admin_menu_kb())
        finally:
            for p in (users_path, words_path, stats_path):
                try:
                    if os.path.exists(p):
                        os.remove(p)
                except Exception:
                    pass
        return

    # === BACKUP ===
    if data == "admin:backup":
        await q.edit_message_text(
            "💾 Creating backup... Please wait.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Cancel", callback_data="admin:main")]])
        )
        try:
            backup_file = create_full_backup()
            if backup_file:
                backup_size = os.path.getsize(backup_file) / (1024 * 1024)
                msg = f"✅ Backup created successfully!\n"
                msg += f"📦 File: {os.path.basename(backup_file)}\n"
                msg += f"📊 Size: {backup_size:.2f} MB\n"
                msg += f"🕐 Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                msg += get_backup_size_info()
                await q.edit_message_text(msg, reply_markup=admin_menu_kb())
                
                # Send backup file to admin
                chat_id = q.message.chat.id
                with open(backup_file, 'rb') as f:
                    await context.bot.send_document(
                        chat_id=chat_id,
                        document=f,
                        filename=os.path.basename(backup_file),
                        caption="💾 Full backup file"
                    )
            else:
                await q.edit_message_text("❌ Error creating backup", reply_markup=admin_menu_kb())
        except Exception as e:
            log.error(f"Backup error: {e}")
            await q.edit_message_text(f"❌ Backup failed: {e}", reply_markup=admin_menu_kb())
        return

    # === RESTORE ===
    if data == "admin:restore":
        backups = list_backups()
        if not backups:
            await q.edit_message_text(
                "❌ No backups available",
                reply_markup=admin_menu_kb()
            )
            return
        
        buttons = []
        for filename, size_mb, ts in backups[:5]:  # Show last 5 backups
            btn_text = f"💾 {ts} ({size_mb:.1f}MB)"
            buttons.append([InlineKeyboardButton(btn_text, callback_data=f"admin:restore_confirm:{filename}")])
        buttons.append([InlineKeyboardButton("⬅️ Cancel", callback_data="admin:main")])
        
        msg = "📋 Available backups:\n\n"
        msg += "\n".join([f"• {f[0]} ({f[1]:.2f}MB)" for f in backups[:5]])
        await q.edit_message_text(msg, reply_markup=InlineKeyboardMarkup(buttons))
        return
    
    # === RESTORE CONFIRM ===
    if data.startswith("admin:restore_confirm:"):
        filename = data.split(":", 2)[2]
        backup_file = os.path.join(os.getenv("BACKUP_DIR", "/tmp/wordl_backups"), filename)
        
        await q.edit_message_text(
            "⚠️ Confirming restore operation...\n"
            "This will overwrite current data with backup data.\n"
            "Current data will be automatically backed up.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Restore", callback_data=f"admin:restore_execute:{filename}"),
                 InlineKeyboardButton("❌ Cancel", callback_data="admin:restore")]
            ])
        )
        return
    
    # === RESTORE EXECUTE ===
    if data.startswith("admin:restore_execute:"):
        filename = data.split(":", 2)[2]
        backup_file = os.path.join(os.getenv("BACKUP_DIR", "/tmp/wordl_backups"), filename)
        
        await q.edit_message_text(
            "🔄 Restoring from backup... Please wait.\n\n"
            "⚠️ Bot will be restarting after restore.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Admin menu", callback_data="admin:main")]])
        )
        
        try:
            success, message = restore_full_backup(backup_file)
            if success:
                await q.edit_message_text(
                    f"✅ Restore completed!\n\n{message}\n\n"
                    f"⚠️ Please restart the bot to apply changes.",
                    reply_markup=admin_menu_kb()
                )
                log.info("Restore completed successfully")
            else:
                await q.edit_message_text(
                    f"{message}\n\nRestoration cancelled.",
                    reply_markup=admin_menu_kb()
                )
        except Exception as e:
            log.error(f"Restore error: {e}")
            await q.edit_message_text(
                f"❌ Restore error: {e}",
                reply_markup=admin_menu_kb()
            )
        return

    # === DELETE BACKUPS ===
    if data == "admin:delete_backups":
        backups = list_backups()
        if not backups:
            await q.edit_message_text(
                "❌ No backups available to delete",
                reply_markup=admin_menu_kb()
            )
            return
        
        buttons = []
        for filename, size_mb, ts in backups[:10]:  # Show last 10 backups
            btn_text = f"🗑️ {ts} ({size_mb:.1f}MB)"
            buttons.append([InlineKeyboardButton(btn_text, callback_data=f"admin:delete_backup_confirm:{filename}")])
        buttons.append([InlineKeyboardButton("⬅️ Cancel", callback_data="admin:main")])
        
        msg = "📋 Select backup to delete:\n\n"
        msg += "\n".join([f"• {f[0]} ({f[1]:.2f}MB)" for f in backups[:10]])
        await q.edit_message_text(msg, reply_markup=InlineKeyboardMarkup(buttons))
        return
    
    # === DELETE BACKUP CONFIRM ===
    if data.startswith("admin:delete_backup_confirm:"):
        filename = data.split(":", 2)[2]
        backup_file = os.path.join(os.getenv("BACKUP_DIR", "/tmp/wordl_backups"), filename)
        
        await q.edit_message_text(
            f"⚠️ Delete backup: {filename}?\n\n"
            "This action cannot be undone!",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Delete", callback_data=f"admin:delete_backup_execute:{filename}"),
                 InlineKeyboardButton("❌ Cancel", callback_data="admin:delete_backups")]
            ])
        )
        return
    
    # === DELETE BACKUP EXECUTE ===
    if data.startswith("admin:delete_backup_execute:"):
        filename = data.split(":", 2)[2]
        backup_file = os.path.join(os.getenv("BACKUP_DIR", "/tmp/wordl_backups"), filename)
        
        try:
            if os.path.exists(backup_file):
                os.remove(backup_file)
                await q.edit_message_text(
                    f"✅ Backup deleted: {filename}",
                    reply_markup=admin_menu_kb()
                )
                log.info(f"Backup deleted: {filename}")
            else:
                await q.edit_message_text(
                    "❌ Backup file not found",
                    reply_markup=admin_menu_kb()
                )
        except Exception as e:
            log.error(f"Delete backup error: {e}")
            await q.edit_message_text(
                f"❌ Delete failed: {e}",
                reply_markup=admin_menu_kb()
            )
        return

    # === GROUPS SECTION ===
    if data == "admin:groups":
        await q.edit_message_text("Groups menu:", reply_markup=admin_groups_submenu_kb())
        return

    # Admin: manage all groups (list with per-group actions)
    if data == "admin:manage_groups":
        # list first page of groups (show id + name)
        rows = fetch_groups_page(0)
        if not rows:
            await q.edit_message_text("No groups.", reply_markup=admin_groups_submenu_kb())
            return
        buttons = []
        for r in rows:
            buttons.append([InlineKeyboardButton(f"{r['name']} (ID:{r['id']})", callback_data=f"admin:group_actions:{r['id']}")])
        buttons.append([InlineKeyboardButton("⬅️ Back", callback_data="admin:groups")])
        await q.edit_message_text("Select group to manage:", reply_markup=InlineKeyboardMarkup(buttons))
        return

    if data.startswith("admin:group_actions:"):
        parts = data.split(":")
        try:
            group_id = int(parts[2])
        except Exception:
            await q.answer("Invalid group id", show_alert=True)
            return
        # show actions that reuse existing callbacks (those callbacks will accept admins)
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("✏️ Rename", callback_data=f"group_rename_select:{group_id}")],
            [InlineKeyboardButton("👥 Add user", callback_data=f"group_add_user_select:{group_id}")],
            [InlineKeyboardButton("➕ Add word", callback_data=f"group_add_select:{group_id}")],
            [InlineKeyboardButton("📚 View words", callback_data=f"admin:group_words:{group_id}:0")],
            [InlineKeyboardButton("🗑️ Delete", callback_data=f"group_delete_select:{group_id}")],
            [InlineKeyboardButton("⬅️ Back", callback_data="admin:manage_groups")]
        ])
        await q.edit_message_text(f"Manage group ID {group_id}", reply_markup=kb)
        return

    if data.startswith("admin:group_words:"):
        # reuse admin group words listing (paginated)
        parts = data.split(":")
        group_id = int(parts[2])
        offset = int(parts[3]) if len(parts) > 3 else 0
        with db() as conn:
            total = conn.execute("SELECT COUNT(*) as cnt FROM words WHERE group_id=?", (group_id,)).fetchone()["cnt"]
            rows = conn.execute("SELECT * FROM words WHERE group_id=? LIMIT 50 OFFSET ?", (group_id, offset)).fetchall()
        if not rows:
            await q.edit_message_text("No words in this group.", reply_markup=admin_groups_submenu_kb())
            return
        text = f"📚 Words in group (showing {offset+1}-{min(offset+50, total)}/{total}):\n```ID     English          Uzbek            Level\n" + ("-"*50) + "\n"
        for w in rows:
            text += f"{w['id']:<6} {w['english']:<17} {w['uzbek']:<17} {w['review_level']}\n"
        text += "```"
        buttons = []
        nav_row = []
        if offset > 0:
            nav_row.append(InlineKeyboardButton("⬅️ Prev", callback_data=f"admin:group_words:{group_id}:{max(offset-50,0)}"))
        if offset + 50 < total:
            nav_row.append(InlineKeyboardButton("Next ➡️", callback_data=f"admin:group_words:{group_id}:{offset+50}"))
        if nav_row:
            buttons.append(nav_row)
        buttons.append([InlineKeyboardButton("⬅️ Back", callback_data=f"admin:group_actions:{group_id}")])
        await q.edit_message_text(text, reply_markup=InlineKeyboardMarkup(buttons))
        return

    if data.startswith("admin:groups:"):
        offset = int(data.split(":")[2]) if len(data.split(":")) > 2 else 0
        total = count_groups()
        text = groups_page_text(offset, total)
        kb = groups_page_kb(offset, total)
        await q.edit_message_text(text, reply_markup=kb)
        return

    if data == "admin:create_group":
        context.user_data["admin_mode"] = "create_group"
        await q.edit_message_text(
            "Send group name (e.g., English A1)",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")]])
        )
        return

    if data == "admin:group_add_word":
        context.user_data["admin_mode"] = "group_add_word"
        await q.edit_message_text(
            "Send group ID and word: group_id:english - translation\nExample: 1:Hello - Hello",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")]])
        )
        return

    await q.edit_message_text("Unknown admin action.", reply_markup=admin_menu_kb())
    return



# =====================
# Poll answer handler (quiz polls)
# =====================

async def on_poll_answer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ans = update.poll_answer
    poll_id = ans.poll_id
    chosen = ans.option_ids[0] if ans.option_ids else None
    info = ACTIVE_POLLS.get(poll_id)
    if not info:
        return
    if ans.user.id != info["tg_user_id"]:
        return

    word_id = info["word_id"]
    db_user_id = info["db_user_id"]
    chat_id = info["chat_id"]
    message_id = info["message_id"]
    correct_idx = info["correct_idx"]
    is_blitz = info.get("is_blitz", False)
    group_id = info.get("group_id")
    question_num = info.get("question_num", 1)

    if chosen is None:
        return

    if chosen == correct_idx:
        record_stat(db_user_id, "correct", word_id, is_blitz=is_blitz, group_id=group_id)
        new_correct_count = QUIZ_SESSIONS.get(ans.user.id, {}).get("correct_count", 0) + 1 if not is_blitz else 0
        if is_blitz:
            sess = BLITZ_SESSIONS.get(ans.user.id)
            if sess:
                sess["correct"] = sess.get("correct", 0) + 1
        if not is_blitz:
            sess = QUIZ_SESSIONS.get(ans.user.id)
            if sess:
                sess["correct_count"] = new_correct_count
    else:
        record_stat(db_user_id, "wrong", word_id, group_id=group_id)
        new_correct_count = QUIZ_SESSIONS.get(ans.user.id, {}).get("correct_count", 0)
        if is_blitz:
            sess = BLITZ_SESSIONS.get(ans.user.id)
            if sess:
                sess["wrong"] = sess.get("wrong", 0) + 1

    try:
        await context.bot.stop_poll(chat_id=chat_id, message_id=message_id)
        await asyncio.sleep(0.5)
    except Exception as e:
        log.warning("stop_poll failed: %s", e)

    if is_blitz:
        await send_blitz_poll_app(context.application, chat_id, db_user_id, info["tg_user_id"], group_id=info["group_id"])
    else:
        next_question_num = question_num + 1
        await send_quiz_poll(context, chat_id, db_user_id, info["tg_user_id"], group_id=info["group_id"], question_num=next_question_num)

    ACTIVE_POLLS.pop(poll_id, None)

# =====================
# Leaderboard handlers
# =====================

def get_leaderboard(period: str) -> list[sqlite3.Row]:
    with db() as conn:
        if period == "daily":
            today = local_date()
            rows = conn.execute("SELECT u.tg_id, u.username, u.points FROM users u WHERE u.active=1 ORDER BY u.points DESC LIMIT 10").fetchall()
        elif period == "weekly":
            start = (local_today() - timedelta(days=6)).isoformat()
            rows = conn.execute("""
                SELECT u.tg_id, u.username, SUM(CASE action WHEN 'correct' THEN 5 WHEN 'wrong' THEN -4 ELSE 0 END) AS points
                FROM stats s JOIN users u ON u.id = s.user_id
                WHERE s.local_date >= ? GROUP BY s.user_id ORDER BY points DESC LIMIT 10
            """, (start,)).fetchall()
        elif period == "monthly":
            ym = local_date()[:7]
            rows = conn.execute("""
                SELECT u.tg_id, u.username, SUM(CASE action WHEN 'correct' THEN 5 WHEN 'wrong' THEN -4 ELSE 0 END) AS points
                FROM stats s JOIN users u ON u.id = s.user_id
                WHERE substr(s.local_date,1,7)=? GROUP BY s.user_id ORDER BY points DESC LIMIT 10
            """, (ym,)).fetchall()
    return rows

async def leader_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = get_or_create_user(update.effective_user.id, update.effective_user.username)
    text = LANGS[get_ui_lang(uid)]["leader_choose"]
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Daily", callback_data="leader:daily"),
         InlineKeyboardButton("📆 Weekly", callback_data="leader:weekly"),
         InlineKeyboardButton("🗓 Monthly", callback_data="leader:monthly")]
    ])
    await update.message.reply_text(text, reply_markup=kb)

async def leader_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    period = q.data.split(":")[1]
    rows = get_leaderboard(period)
    if not rows:
        await q.edit_message_text(LANGS[get_ui_lang(get_or_create_user(q.from_user.id, q.from_user.username))]["leader_none"])
        return
    title = {"daily":"📅 Daily", "weekly":"📆 Weekly", "monthly":"🗓 Monthly"}[period]
    text = f"{title} leaderboard:\n\n"
    for i,r in enumerate(rows,1):
        uname = f"@{r['username']}" if r['username'] else str(r['tg_id'])
        text += f"{i}. {uname} — {r['points']} points\n"
    await q.edit_message_text(text, reply_markup=q.message.reply_markup)

async def noop_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

async def grammar_file_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle grammar file selection."""
    from grammar import handle_grammar_file_selection
    await handle_grammar_file_selection(update, context)

async def grammar_pagination_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle grammar pagination (Next/Previous)."""
    from grammar import handle_grammar_pagination
    await handle_grammar_pagination(update, context)

async def ielts_file_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle IELTS book selection."""
    from ielts import handle_cambridge_book_selection
    await handle_cambridge_book_selection(update, context)

async def ielts_pagination_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle IELTS send (book or test)."""
    from ielts import handle_ielts_send
    await handle_ielts_send(update, context)

async def ielts_back_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle IELTS back button."""
    from ielts import handle_ielts_back
    await handle_ielts_back(update, context)

async def group_rename_select_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    group_id = int(q.data.split(":")[1])
    if not (is_group_owner(uid, group_id) or is_admin(uid)):
        await q.edit_message_text("No permission.")
        return
    context.user_data["pending_group_rename"] = group_id
    await q.edit_message_text(L["group_rename_prompt"])

async def group_delete_select_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    group_id = int(q.data.split(":")[1])
    if not (is_group_owner(uid, group_id) or is_admin(uid)):
        await q.edit_message_text("No permission.")
        return
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(L["group_delete_yes"], callback_data=f"group_delete_confirm:{group_id}:yes"),
         InlineKeyboardButton(L["group_delete_no"], callback_data=f"group_delete_confirm:{group_id}:no")]
    ])
    await q.edit_message_text(L["group_delete_confirm"], reply_markup=kb)

async def group_delete_confirm_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    u = q.from_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    parts = q.data.split(":")
    group_id = int(parts[1])
    confirm = parts[2]
    if confirm == "yes":
        ok = delete_group(group_id, uid)
        msg = L["group_delete_success"] if ok else L["group_delete_fail"]
        await q.edit_message_text(msg)
    else:
        await q.edit_message_text("Cancelled.")

async def add_user_to_group_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    context.user_data["awaiting_add_user_to_group"] = True
    await update.message.reply_text(L["add_user_to_group_prompt"])

# =====================
# Math module callbacks handler
# =====================

async def math_callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Route math module callbacks to the appropriate handler"""
    global MATH_HANDLER
    
    if not MATH_HANDLER:
        try:
            await update.callback_query.answer("Math module not available", show_alert=True)
        except Exception:
            pass
        return
    
    query = update.callback_query
    data = query.data
    u = query.from_user
    
    # Store db_user_id in context
    uid = get_or_create_user(u.id, u.username)
    context.user_data['db_user_id'] = uid
    
    try:
        # Route based on callback pattern
        if data == "trig_enter_code":
            await MATH_HANDLER.request_code(update, context)
        elif data == "trig_help":
            await query.answer("Kodi admin bilan oling", show_alert=True)
        elif data.startswith("trig_select_"):
            # Quiz mode selection
            await MATH_HANDLER.send_welcome_message(update, context)
        elif data == "trig_back_main":
            await MATH_HANDLER.send_welcome_message(update, context)
        elif data == "trig_view_values":
            # Show angle selection for values
            await MATH_HANDLER.show_trig_values(update, context)
        elif data.startswith("trig_angle_"):
            # Show values for selected angle
            angle = int(data.split("_")[2])
            await MATH_HANDLER.show_angle_values(update, context, angle)
        elif data == "trig_back_select":
            # Show quiz mode selection
            await query.edit_message_text(
                "📐 Savollar sonini tanlang:",
                reply_markup=MATH_HANDLER.get_quiz_mode_keyboard()
            )
        elif data.startswith("trig_quiz_"):
            num = int(data.split("_")[2])
            await MATH_HANDLER.start_quiz(update, context, num)
        elif data == "trig_next_question":
            await MATH_HANDLER.next_question_handler(update, context)
        elif data.startswith("trig_answer_"):
            answer_idx = int(data.split("_")[2])
            await MATH_HANDLER.process_answer(update, context, answer_idx)
        elif data == "trig_quit_quiz":
            await MATH_HANDLER.finish_quiz(update, context)
        elif data == "trig_view_stats":
            await MATH_HANDLER.show_statistics(update, context)
        elif data == "trig_question_stats":
            await MATH_HANDLER.show_question_statistics(update, context)
        elif data == "trig_leaderboard":
            await MATH_HANDLER.show_leaderboard(update, context)
        else:
            await query.answer("Unknown command", show_alert=True)
    except Exception as e:
        log.error(f"Math callback handler error: {e}")
        try:
            await query.answer(f"❌ Xato: {e}", show_alert=True)
        except Exception:
            pass

# =====================
# Dispatch text messages (add word, admin modes, etc.)
# =====================

async def dispatch_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    if is_banned(u.id):
        return
    uid = get_or_create_user(u.id, u.username)
    L = LANGS.get(get_ui_lang(uid), LANGS["UZ"])
    txt = update.message.text.strip()
    
    # Store db_user_id in context for math module
    context.user_data['db_user_id'] = uid
    
    # If math handler is present and user is in a math flow (waiting_for_code or active quiz),
    # forward the message to the math module to handle answers and codes (non-inline flow)
    if MATH_HANDLER and (context.user_data.get('waiting_for_code') or context.user_data.get('math_quiz_session')):
        try:
            result = await MATH_HANDLER.handle_incoming_message(update, context)
            if result:
                return
        except Exception as e:
            log.error(f"Math module error: {e}")
            try:
                await update.message.reply_text(f"❌ Xato: {e}")
            except Exception:
                pass
            return

    # Check if this is the secret math code (also allow direct code trigger when not in waiting state)
    if txt == "0107" and MATH_HANDLER:
        try:
            result = await MATH_HANDLER.handle_incoming_message(update, context)
            if result:
                return
        except Exception as e:
            log.error(f"Math module error: {e}")
            await update.message.reply_text(f"❌ Xato: {e}")
            return
    
    # Handle large messages and floods with RetryAfter exception
    try:
        msg_len = len(txt)
        if msg_len > 4096:
            await update.message.reply_text(L.get("message_too_long", "Message is too long (max 4096 chars)."))
            return
    except RetryAfter as e:
        log.warning(f"RetryAfter exception: waiting {e.retry_after} seconds")
        await asyncio.sleep(e.retry_after)
        return
    except BadRequest as e:
        log.error(f"BadRequest in dispatch_text: {e}")
        return

    # Qo'shish menyusidan keyin guruh tanlash
    if txt == L["menu_add"]:
        context.user_data["pending_action"] = "add"
        await select_group(update, context)
        return

    # Guruh tanlangan va so'zlar yuborilgan
    if "awaiting_add" in context.user_data:
        group_id = context.user_data["awaiting_add"]
        del context.user_data["awaiting_add"]
        lines = [l for l in (txt or "").strip().split('\n')]
        # Multiple lines -> bulk add
        if len(lines) > 1:
            if group_id:
                try:
                    added_count, errors = add_words_from_lines(uid, lines, group_id=group_id)
                    if added_count > 0:
                        msg = L.get("multiple_added", "{count} words added").format(count=added_count)
                        if errors:
                            msg += "\n" + L.get("errors", "Errors:") + " " + "\n".join(errors[:5])
                        await update.message.reply_text(msg)
                    else:
                        await update.message.reply_text(L["format_error"])
                except RetryAfter as e:
                    log.warning(f"RetryAfter in bulk add: {e.retry_after}s")
                    await update.message.reply_text(L.get("too_many_requests", "Too many requests. Please try again later."))
                except Exception as e:
                    log.error(f"Error in bulk add: {e}")
                    await update.message.reply_text(L.get("error_occurred", "An error occurred. Please try again."))
                return
            else:
                # need to create a group name first
                context.user_data["awaiting_group_name_for_multi"] = txt
                await update.message.reply_text(t_for(uid, "group_name_prompt_for_multi"))
                return

        # Single line -> parse and add
        try:
            eng, uz = parse_word_line(txt)
            add_word(uid, eng, uz, group_id=group_id)
            await update.message.reply_text(L["added_ok"].format(eng=eng, uz=uz))
        except RetryAfter as e:
            log.warning(f"RetryAfter in single add: {e.retry_after}s")
            await update.message.reply_text(L.get("too_many_requests", "Too many requests. Please try again later."))
        except BadRequest as e:
            log.error(f"BadRequest in single add: {e}")
            await update.message.reply_text(L.get("error_occurred", "An error occurred. Please try again."))
        except Exception as e:
            log.warning("add failed: %s", e)
            await update.message.reply_text(L["format_error"])
        return

    # Yangi guruh nomi yuborilganda va unga ko'p so'z qo'shilganda
    if "awaiting_group_name_for_multi" in context.user_data:
        multi_txt = context.user_data["awaiting_group_name_for_multi"]
        del context.user_data["awaiting_group_name_for_multi"]
        gid = create_group(txt, uid)
        lines = [l for l in (multi_txt or "").strip().split('\n')]
        added_count, errors = add_words_from_lines(uid, lines, group_id=gid)
        if added_count > 0:
            msg = L["multiple_added"].format(count=added_count)
            if errors:
                msg += "\n" + L["errors"] + " " + "\n".join(errors)
            await update.message.reply_text(msg)
        else:
            await update.message.reply_text(L["format_error"])
        return

    # Quiz menyusi
    if txt == L["menu_quiz"]:
        context.user_data["pending_action"] = "quiz"
        await select_group(update, context)
        return

    # Statistikalar
    if txt == L["menu_stats"]:
        months = month_list_for_user(uid)
        if not months:
            await update.message.reply_text(L["no_stats"])
            return
        await update.message.reply_text("Choose month (YYYY-MM):", reply_markup=month_keyboard(months))
        return

    # So'zlar
    if txt == L["menu_words"]:
        context.user_data["pending_action"] = "words"
        await select_group(update, context)
        return

    # Eslatma paneli
    if txt == L["menu_remind"]:
        await open_reminder_panel(update, context)
        return

    # Import/export
    if txt == L["menu_io"]:
        await group_io_command(update, context)
        return

    # Blitz rejimi
    if txt == L["menu_blitz"]:
        context.user_data["pending_action"] = "blitz"
        await select_group(update, context)
        return

    # Reyting
    if txt == L["menu_leader"]:
        await leader_handler(update, context)
        return

    # Til tanlash
    if txt == L["menu_lang"]:
        await update.message.reply_text(L["choose_lang"], reply_markup=language_keyboard())
        return

    # Guruhlar
    if txt == L["menu_groups"]:
        groups = get_user_groups(uid)
        buttons = []
        if not groups:
            buttons.append([InlineKeyboardButton("➕ Yangi guruh yaratish", callback_data="create_group_inline")])
            await update.message.reply_text("No groups.", reply_markup=InlineKeyboardMarkup(buttons))
            return
        for g in groups:
            row = [
                InlineKeyboardButton(g["name"], callback_data=f"group_select:{g['id']}"),
                InlineKeyboardButton("✏️", callback_data=f"group_rename_select:{g['id']}"),
                InlineKeyboardButton("🗑", callback_data=f"group_delete_select:{g['id']}"),
                InlineKeyboardButton("📁", callback_data=f"group_io_select:{g['id']}:menu"),    
                InlineKeyboardButton("➕", callback_data=f"group_add_select:{g['id']}"),
                InlineKeyboardButton("👤 Add user", callback_data=f"group_add_user_select:{g['id']}")
            ]
            buttons.append(row)
        buttons.append([InlineKeyboardButton("➕ Yangi guruh yaratish", callback_data="create_group_inline")])
        await update.message.reply_text("Groups:", reply_markup=InlineKeyboardMarkup(buttons))
        return

    # Grammatika
    if txt == L["menu_grammar"]:
        from grammar import show_grammar_files
        await show_grammar_files(update, context)
        return

    # IELTS
    if txt == L["menu_ielts"]:
        from ielts import show_cambridge_books
        await show_cambridge_books(update, context)
        return

    # Math / Trigonometry
    if txt == L["menu_math"]:
        if MATH_HANDLER:
            # Prompt the user to enter the secret code when they press Math
            context.user_data['waiting_for_code'] = True
            try:
                await update.message.reply_text(
                    "🔐 Maxfiy bo'limni ochish uchun kodni kiriting."
                )
            except Exception:
                # Fallback to handler request if message reply fails
                try:
                    await MATH_HANDLER.request_code(update, context)
                except Exception:
                    pass
        else:
            await update.message.reply_text("❌ Math module is not available")
        return

    # Settings
    if txt == L["menu_settings"]:
        await open_settings_panel(update, context)
        return

    # Admin menyu
    if txt == L["menu_admin"]:
        if is_admin(u.id):
            await open_admin_panel(update, context)
        return

    # Duel, Hunt, Share, Progress features removed from this build.
    # Menu entries have been removed from the keyboard, so no handler is needed here.

    # Guruh yaratish
    if context.user_data.get("pending_group_create"):
        del context.user_data["pending_group_create"]
        gid = create_group(txt, uid)
        await update.message.reply_text(f"Group created: {txt} (ID: {gid})")
        return

    # Guruh nomini o'zgartirish
    if "pending_group_rename" in context.user_data:
        group_id = context.user_data["pending_group_rename"]
        del context.user_data["pending_group_rename"]
        ok = rename_group(group_id, txt, uid)
        msg = "Group name changed." if ok else "Group name not changed (ownership or error)."
        await update.message.reply_text(msg)
        return

    # Maxsus vaqt
    if context.user_data.get("awaiting_custom_time"):
        del context.user_data["awaiting_custom_time"]
        try:
            _parse_hhmm(txt)
            set_settings(uid, remind_time=txt)
            schedule_user_reminder(context.application, u.id, txt, get_settings(uid)["remind_enabled"])
            await update.message.reply_text(L["time_changed"])
        except Exception:
            await update.message.reply_text(L["invalid_time_format"])
        return

    # Maxsus maqsad
    if context.user_data.get("awaiting_custom_goal"):
        del context.user_data["awaiting_custom_goal"]
        try:
            goal = int(txt.strip())
            if goal < 1 or goal > 100:
                await update.message.reply_text(L["goal_range_error"])
                return
            set_settings(uid, daily_goal=goal)
            await update.message.reply_text(f"Goal changed to {goal} words/day.")
        except ValueError:
            await update.message.reply_text(L["invalid_format"])
        return

    # Custom quiz repeat
    if context.user_data.get("awaiting_quiz_repeat"):
        del context.user_data["awaiting_quiz_repeat"]
        try:
            v = int(txt.strip())
            if v < 1 or v > 100:
                await update.message.reply_text(L["invalid_number"])
                return
            set_settings(uid, quiz_repeat=v)
            await update.message.reply_text(L["settings_changed"])
        except ValueError:
            await update.message.reply_text(L["invalid_number"])
        return

    # Custom restart on incorrect
    if context.user_data.get("awaiting_restart_incorrect"):
        del context.user_data["awaiting_restart_incorrect"]
        try:
            v = int(txt.strip())
            if v < 1 or v > 100:
                await update.message.reply_text(L["invalid_number"])
                return
            set_settings(uid, restart_on_incorrect=v)
            await update.message.reply_text(L["settings_changed"])
        except ValueError:
            await update.message.reply_text(L["invalid_number"])
        return

    # Guruhga so'z qo'shish
    if "awaiting_group_add" in context.user_data:
        group_id = context.user_data["awaiting_group_add"]
        del context.user_data["awaiting_group_add"]
        try:
            eng, uz = parse_word_line(txt)
            add_word(uid, eng, uz, group_id=group_id)
            await update.message.reply_text(L["group_add_success"].format(eng=eng, uz=uz))
        except Exception:
            await update.message.reply_text(L["format_error"])
        return

    # Guruhga user qo'shish
    if "awaiting_add_user_to_group" in context.user_data:
        del context.user_data["awaiting_add_user_to_group"]
        try:
            user_tg_id = int(txt.strip())
            user_id = get_or_create_user(user_tg_id, None)
            ok = add_user_to_group(user_id, context.user_data.get("awaiting_add_user_to_group"), uid)
            msg = L["user_added_to_group"] if ok else L["no_permission"]
            await update.message.reply_text(msg)
        except Exception as e:
            await update.message.reply_text(f"Error: {e}")
        return

    # Guruhga user qo'shish (callback orqali)
    if "awaiting_add_user_to_group_from_cb" in context.user_data:
        group_id = context.user_data["awaiting_add_user_to_group_from_cb"]
        del context.user_data["awaiting_add_user_to_group_from_cb"]
        try:
            user_tg_id = int(txt.strip())
            user_id = get_or_create_user(user_tg_id, None)
            ok = add_user_to_group(user_id, group_id, uid)
            msg = L["user_added_to_group"] if ok else L["no_permission"]
            await update.message.reply_text(msg)
        except Exception as e:
            await update.message.reply_text(f"Error: {e}")
        return

    # So'z o'chirish rejimi - raqam qilgan
    if context.user_data.get("awaiting_delete_number"):
        del context.user_data["awaiting_delete_number"]
        delete_data = context.user_data.pop("delete_mode_data", None)
        if not delete_data:
            await update.message.reply_text(L["not_found_or_no_permission"])
            return
        try:
            word_num = int(txt.strip())
            if word_num < 1 or word_num > len(delete_data["rows"]):
                await update.message.reply_text(L["not_found_or_no_permission"])
                return
            selected_word = delete_data["rows"][word_num - 1]
            ok = delete_word_if_owner(selected_word["id"], delete_data["uid"])
            if ok:
                await update.message.reply_text(L["deleted"])
                # Send the words page again to refresh
                tg_id = delete_data["tg_id"]
                offset = delete_data["offset"]
                days = delete_data["days"]
                group_id = delete_data["group_id"]
                # Create a mock callback query context to reuse send_words_page
                q = update.message
                # We'll send the list directly instead
                new_rows = fetch_words_page(delete_data["uid"], offset, days=days, group_id=group_id)
                if new_rows:
                    lines = [f"{i+1}. {r['english']} — {r['uzbek']}" for i,r in enumerate(new_rows)]
                    body = "🗑 " + L["delete_mode"] + "\n" + "\n".join(lines) + "\n\n" + L["awaiting_delete_number"]
                    nav = [[InlineKeyboardButton("⬅️ Back", callback_data=f"w:{tg_id}:{offset}:{days or 'all'}:{group_id or 'none'}"), InlineKeyboardButton("Close", callback_data=f"w:{tg_id}:close")]]
                    context.user_data["delete_mode_data"] = {
                        "tg_id": tg_id,
                        "offset": offset,
                        "days": days,
                        "group_id": group_id,
                        "rows": new_rows,
                        "uid": delete_data["uid"]
                    }
                    await update.message.reply_text(body, reply_markup=InlineKeyboardMarkup(nav))
                    context.user_data["awaiting_delete_number"] = True
                else:
                    await update.message.reply_text(L["no_words"])
            else:
                await update.message.reply_text(L["not_found_or_no_permission"])
        except ValueError:
            await update.message.reply_text(L["not_found_or_no_permission"])
        return

    # Admin rejimlari
    if "admin_mode" in context.user_data:
        mode = context.user_data["admin_mode"]
        if mode == "bc_wait_text":
            context.user_data["bc_text"] = txt
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Send", callback_data="admin:bc_send"),
                 InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")]
            ])
            await update.message.reply_text(f"Broadcast: {txt}\nConfirm:", reply_markup=kb)
            return
        elif mode == "user_add_word":
            try:
                parts = txt.split(":", 1)
                target_uid = int(parts[0])
                rest = parts[1].strip()
                eng, uz = parse_word_line(rest)
                admin_add_word_to_user(target_uid, eng, uz)
                await update.message.reply_text("Word added.")
            except Exception as e:
                await update.message.reply_text(f"Error: {e}")
            finally:
                del context.user_data["admin_mode"]
            return
        elif mode == "user_edit_points":
            try:
                parts = txt.split(":", 1)
                target_uid = int(parts[0])
                points = int(parts[1].strip())
                set_user_points(target_uid, points)
                await update.message.reply_text("Points edited.")
            except Exception as e:
                await update.message.reply_text(f"Error: {e}")
            finally:
                del context.user_data["admin_mode"]
            return
        elif mode == "create_group":
            try:
                gid = create_group(txt, uid)
                await update.message.reply_text(f"Group created: {txt} (ID: {gid})")
            except Exception as e:
                await update.message.reply_text(f"Error: {e}")
            finally:
                del context.user_data["admin_mode"]
            return
        elif mode == "group_add_word":
            try:
                parts = txt.split(":", 1)
                group_id = int(parts[0])
                rest = parts[1].strip()
                eng, uz = parse_word_line(rest)
                add_word(uid, eng, uz, group_id=group_id)
                await update.message.reply_text("Word added to group.")
            except Exception as e:
                await update.message.reply_text(f"Error: {e}")
            finally:
                del context.user_data["admin_mode"]
            return

    # Admin qidiruvi
    if context.user_data.get("admin_await_user"):
        del context.user_data["admin_await_user"]
        db_id = get_user_db_id_from_query(txt)
        if not db_id:
            await update.message.reply_text("Not found.", reply_markup=admin_menu_kb())
            return
        text = get_user_info_text(db_id)
        kb = get_user_info_kb(db_id)
        await update.message.reply_text(text, reply_markup=kb)
        return

    # Oddiy qo'shish (agar matnda "-" bo'lsa)
    if "-" in txt or "–" in txt or "—" in txt:
        try:
            normalized = txt.replace("–", "-").replace("—", "-")
            eng, uz = normalized.split("-", 1)
            eng = eng.strip()
            uz = uz.strip()
            add_word(uid, eng, uz)
            await update.message.reply_text(L["added_ok"].format(eng=eng, uz=uz))
        except Exception as e:
            log.warning("add failed: %s", e)
            await update.message.reply_text(L["format_error"])
    else:
        await update.message.reply_text("Unknown command. Use the menu.")
# =====================
# Scheduled Backup Job
# =====================

async def scheduled_backup_job(context: ContextTypes.DEFAULT_TYPE):
    """Runs weekly to create automatic backup of the database and files."""
    try:
        backup_file = create_full_backup()
        backup_size_info = get_backup_size_info()
        log.info(f"✅ Automated weekly backup created: {backup_file}")
        log.info(f"📊 Backup storage info: {backup_size_info}")
        
        # Optionally send notification to admins (if admin list is available)
        # For now, just log the successful backup
        
    except Exception as e:
        log.error(f"❌ Scheduled backup failed: {e}")

# =====================
# Entry point: main()
# =====================

def main():
    global MATH_HANDLER
    init_db()
    # Bot token must be set to run polling
    if not BOT_TOKEN:
        log.error("BOT_TOKEN is not set. Exiting main().")
        return

    request = HTTPXRequest(read_timeout=30, connect_timeout=30, write_timeout=30)
    app = ApplicationBuilder().token(BOT_TOKEN).request(request).job_queue(JobQueue()).build()

    # Initialize math handler
    if MATH_AVAILABLE:
        try:
            MATH_HANDLER = MathBotHandler(DB_PATH)
            log.info("✅ Math module initialized successfully")
        except Exception as e:
            log.warning(f"Failed to initialize math module: {e}")
            MATH_HANDLER = None

    try:
        reschedule_all(app)
    except Exception as e:
        log.warning("reschedule_all failed: %s", e)
    
    # Automatic backups disabled - backups are now manual only (triggered by admin)
    # try:
    #     app.job_queue.run_repeating(scheduled_backup_job, interval=604800, first=60)
    #     log.info("✅ Weekly backup scheduler enabled (runs every 7 days)")
    # except Exception as e:
    #     log.error(f"Failed to schedule weekly backup job: {e}")

    # === Komandalar ===
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("create_group", create_group_command))
    app.add_handler(CommandHandler("rename_group", rename_group_command))
    app.add_handler(CommandHandler("delete_group", delete_group_command))
    app.add_handler(CommandHandler("group_io", group_io_command))
    app.add_handler(CommandHandler("add_user_to_group", add_user_to_group_command))

    # === Callbacklar ===
    app.add_handler(CallbackQueryHandler(set_language_cb, pattern="^lang:"))
    app.add_handler(CallbackQueryHandler(group_select_cb, pattern="^group_select:"))
    app.add_handler(CallbackQueryHandler(quiz_continue_cb, pattern="^quiz_continue:"))
    app.add_handler(CallbackQueryHandler(group_rename_select_cb, pattern="^group_rename_select:"))
    app.add_handler(CallbackQueryHandler(group_delete_select_cb, pattern="^group_delete_select:"))
    app.add_handler(CallbackQueryHandler(group_delete_confirm_cb, pattern="^group_delete_confirm:"))
    app.add_handler(CallbackQueryHandler(group_add_select_cb, pattern="^group_add_select:"))
    app.add_handler(CallbackQueryHandler(group_add_user_select_cb, pattern="^group_add_user_select:"))
    app.add_handler(CallbackQueryHandler(create_group_inline_cb, pattern="^create_group_inline$"))
    
    # Math module callbacks
    if MATH_HANDLER:
        app.add_handler(CallbackQueryHandler(math_callback_handler, pattern="^trig_"))
    app.add_handler(CallbackQueryHandler(grammar_file_cb, pattern="^grammar_file:"))
    app.add_handler(CallbackQueryHandler(grammar_pagination_cb, pattern="^grammar_page:"))
    app.add_handler(CallbackQueryHandler(ielts_file_cb, pattern="^ielts_book:"))
    app.add_handler(CallbackQueryHandler(ielts_pagination_cb, pattern="^ielts_send:"))
    app.add_handler(CallbackQueryHandler(ielts_back_cb, pattern="^ielts_back$"))

    app.add_handler(MessageHandler(filters.Document.ALL, on_document))
    app.add_handler(PollAnswerHandler(on_poll_answer))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, dispatch_text))

    app.add_handler(CallbackQueryHandler(stats_cb, pattern="^stats_"))
    app.add_handler(CallbackQueryHandler(words_cb, pattern=r"^(w:|wd:|wdx:|wf:|wfr:|io:|group_io_select:|group_io_back|group_io_cancel|group_io:|wclear:|wclear_confirm:)"))
    app.add_handler(CallbackQueryHandler(admin_words_cb, pattern=r"^admin:(words|wf|wfr|wd|wdx):"))
    app.add_handler(CallbackQueryHandler(reminder_cb, pattern="^rem:"))
    app.add_handler(CallbackQueryHandler(settings_cb, pattern="^settings:"))
    app.add_handler(CallbackQueryHandler(admin_cb, pattern="^admin:"))
    app.add_handler(CallbackQueryHandler(leader_cb, pattern="^leader:"))
    app.add_handler(CallbackQueryHandler(blitz_start_cb, pattern="^blitz_start:"))
    app.add_handler(CallbackQueryHandler(import_cancel_cb, pattern="^import_cancel$"))
    app.add_handler(CallbackQueryHandler(noop_cb, pattern="^noop$"))

    # word1 handlers were removed; no external registration.

    log.info("Bot is running. Press Ctrl+C to stop.")
    try:
        app.run_polling()
    except Exception as e:
        log.exception("app.run_polling crashed: %s", e)

if __name__ == "__main__":
    main()
